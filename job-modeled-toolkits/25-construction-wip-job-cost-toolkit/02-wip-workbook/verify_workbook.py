"""Checks the WIP workbook against the engine's schedule, to the cent.

The workbook holds live formulas, not baked-in numbers, so this tool proves two
things for every job. First, that each derived cell holds exactly the formula it
should. Second, that the formula, computed straight from the workbook's own input
cells, equals the engine's figure to the cent. The computation runs through the
small evaluator in formula_eval.py, never through Excel and never through the
engine code, so it is an independent check.

It also confirms the totals row and the Dashboard sheet add up to the schedule.

Usage:
  python verify_workbook.py
  python verify_workbook.py --schedule wip_schedule.csv --workbook wip_workbook.xlsx

Prints PASS or FAIL for each section and exits non-zero if anything disagrees.
"""

import argparse
import csv
from decimal import Decimal, ROUND_HALF_UP

from openpyxl import load_workbook

import formulas
from formula_eval import Evaluator

CENT = Decimal("0.01")
RATIO = Decimal("0.0001")


def money(value):
    return Decimal(value).quantize(CENT, rounding=ROUND_HALF_UP)


def ratio(value):
    return Decimal(value).quantize(RATIO, rounding=ROUND_HALF_UP)


def read_schedule(path):
    with open(path, newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


class Report:
    def __init__(self):
        self.passed = 0
        self.failures = []

    def check(self, ok, message):
        if ok:
            self.passed += 1
        else:
            self.failures.append(message)


def verify(schedule_path, workbook_path):
    rows = read_schedule(schedule_path)
    wb = load_workbook(workbook_path)
    ws = wb["WIP Schedule"]
    ev = Evaluator(wb, "WIP Schedule")
    report = Report()

    headers = [ws["%s%d" % (col, formulas.HEADER_ROW)].value for col, *_ in formulas.COLUMNS]
    report.check(headers == formulas.headers(), "header row does not match the expected columns")

    first = formulas.FIRST_DATA_ROW
    print("Verifying %s against %s" % (workbook_path, schedule_path))
    print()

    for index, row in enumerate(rows):
        r = first + index
        job = row["job_id"]

        for col, field, _header, kind in formulas.COLUMNS:
            coord = "%s%d" % (col, r)
            cell = ws[coord]
            if kind == "text":
                report.check(cell.value == row[field], "%s: %s text mismatch" % (job, col))
            elif kind == "input":
                report.check(
                    money(Decimal(str(cell.value))) == money(row[field]),
                    "%s: %s input mismatch" % (job, col),
                )
            else:
                expected_formula = formulas.cell_formula(col, r)
                report.check(
                    cell.value == expected_formula,
                    "%s: %s formula is %r, expected %r" % (job, col, cell.value, expected_formula),
                )
                computed = ev.evaluate(cell.value)
                if col == "M":
                    report.check(computed == row["status"], "%s: status mismatch" % job)
                elif col == "G":
                    report.check(ratio(computed) == ratio(row[field]),
                                 "%s: percent complete mismatch" % job)
                else:
                    report.check(money(computed) == money(row[field]),
                                 "%s: %s value mismatch (%s vs %s)" % (job, col, computed, row[field]))

        earned = ev.evaluate(ws["H%d" % r].value)
        over_under = ev.evaluate(ws["L%d" % r].value)
        print("  %-8s earned %14s   over/under %13s   %s" % (
            job, money(earned), money(over_under), row["status"]))

    # Totals row.
    last = first + len(rows) - 1
    total_row = last + 1
    print()
    expected_totals = _expected_totals(rows)
    for col in ("C", "D", "E", "F", "H", "I", "J", "K", "L"):
        computed = ev.evaluate(ws["%s%d" % (col, total_row)].value)
        report.check(money(computed) == expected_totals[col],
                     "totals: column %s is %s, expected %s" % (col, money(computed), expected_totals[col]))
    print("  Totals row: earned %s, billed %s, net over/under %s" % (
        expected_totals["H"], expected_totals["F"], expected_totals["L"]))

    # Dashboard sheet.
    _verify_dashboard(wb, rows, expected_totals, report)

    print()
    if report.failures:
        print("FAIL: %d check(s) failed." % len(report.failures))
        for message in report.failures:
            print("  - %s" % message)
        return 1
    print("PASS: %d checks. Every live formula reproduces the engine to the cent." % report.passed)
    return 0


def _expected_totals(rows):
    def column_total(field):
        return money(sum((Decimal(row[field]) for row in rows), Decimal("0")))

    return {
        "C": column_total("contract_value"),
        "D": column_total("estimated_total_cost"),
        "E": column_total("cost_to_date"),
        "F": column_total("billed_to_date"),
        "H": column_total("earned_revenue"),
        "I": column_total("cost_to_complete"),
        "J": column_total("estimated_gross_profit"),
        "K": column_total("gross_profit_to_date"),
        "L": column_total("over_under_billing"),
    }


def _verify_dashboard(wb, rows, expected_totals, report):
    ws = wb["Dashboard"]
    ev = Evaluator(wb, "Dashboard")
    first = formulas.FIRST_DATA_ROW
    last = first + len(rows) - 1
    counts = {
        "Underbilled jobs": sum(1 for row in rows if row["status"] == "Underbilled"),
        "Overbilled jobs": sum(1 for row in rows if row["status"] == "Overbilled"),
        "Even jobs": sum(1 for row in rows if row["status"] == "Even"),
    }
    label_to_total = {
        "Total contract value": expected_totals["C"],
        "Total estimated cost": expected_totals["D"],
        "Total cost to date": expected_totals["E"],
        "Total billed to date": expected_totals["F"],
        "Total earned revenue": expected_totals["H"],
        "Net over/under billing": expected_totals["L"],
        "Gross profit to date": expected_totals["K"],
        "Est. gross profit at completion": expected_totals["J"],
    }

    r = 3
    for label, _formula, _kind in formulas.dashboard_rows(first, last):
        value = ev.evaluate(ws["B%d" % r].value)
        if label in label_to_total:
            report.check(money(value) == label_to_total[label],
                         "dashboard: %s is %s, expected %s" % (label, money(value), label_to_total[label]))
        else:
            report.check(int(value) == counts[label],
                         "dashboard: %s is %s, expected %s" % (label, value, counts[label]))
        r += 1


def main(argv=None):
    parser = argparse.ArgumentParser(description="Verify the WIP workbook against the schedule CSV.")
    parser.add_argument("--schedule", default="wip_schedule.csv")
    parser.add_argument("--workbook", default="wip_workbook.xlsx")
    args = parser.parse_args(argv)
    return verify(args.schedule, args.workbook)


if __name__ == "__main__":
    raise SystemExit(main())
