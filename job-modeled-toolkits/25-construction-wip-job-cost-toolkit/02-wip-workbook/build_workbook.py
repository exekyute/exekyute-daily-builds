"""Builds the WIP workbook from the engine's schedule CSV.

Reads wip_schedule.csv (written by the job-cost engine in 01) and writes
wip_workbook.xlsx: a WIP Schedule sheet whose input columns hold the four job
numbers as values and whose derived columns hold live Excel formulas, plus a
Dashboard sheet that totals the schedule. Because the derived cells are formulas,
opening the workbook and changing an input recomputes the row.

Usage:
  python build_workbook.py
  python build_workbook.py --schedule wip_schedule.csv --out wip_workbook.xlsx

Requires openpyxl (see requirements.txt). Every other tool in this toolkit is
standard-library only.
"""

import argparse
import csv

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import formulas

BASE = "1F3A5F"        # navy, the base tone
ACCENT = "DCE6F1"      # light blue, the accent tone
OVERBILLED = "F8D7DA"  # soft red, billings in excess of earned revenue
UNDERBILLED = "D4EDDA"  # soft green, earned revenue in excess of billings

MONEY_FORMAT = '$#,##0.00'
PERCENT_FORMAT = '0.0%'

HEADER_FONT = Font(bold=True, color="FFFFFF")
TOTAL_FONT = Font(bold=True)
HEADER_FILL = PatternFill("solid", fgColor=BASE)
TOTAL_FILL = PatternFill("solid", fgColor=ACCENT)


def read_schedule(path):
    with open(path, newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def write_schedule_sheet(ws, rows):
    ws.title = "WIP Schedule"
    first = formulas.FIRST_DATA_ROW
    last = first + len(rows) - 1
    total_row = last + 1

    for col, _field, header, _kind in formulas.COLUMNS:
        cell = ws["%s%d" % (col, formulas.HEADER_ROW)]
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for index, row in enumerate(rows):
        r = first + index
        for col, field, _header, kind in formulas.COLUMNS:
            cell = ws["%s%d" % (col, r)]
            if kind == "text":
                cell.value = row[field]
            elif kind == "input":
                cell.value = float(row[field])
            else:
                cell.value = formulas.cell_formula(col, r)
            _format_cell(cell, col)

    ws["A%d" % total_row].value = "Total"
    ws["A%d" % total_row].font = TOTAL_FONT
    for col, _field, _header, _kind in formulas.COLUMNS:
        formula = formulas.total_formula(col, first, last)
        if formula is None:
            continue
        cell = ws["%s%d" % (col, total_row)]
        cell.value = formula
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        _format_cell(cell, col)
    for col in ("B", "M"):
        ws["%s%d" % (col, total_row)].fill = TOTAL_FILL

    _add_status_shading(ws, first, last)
    _set_widths(ws)
    ws.freeze_panes = "A%d" % first


def _format_cell(cell, col):
    if col in formulas.MONEY_COLUMNS:
        cell.number_format = MONEY_FORMAT
    elif col in formulas.PERCENT_COLUMNS:
        cell.number_format = PERCENT_FORMAT
        cell.alignment = Alignment(horizontal="center")
    elif col == "M":
        cell.alignment = Alignment(horizontal="center")


def _add_status_shading(ws, first, last):
    span = "L%d:L%d" % (first, last)
    ws.conditional_formatting.add(
        span,
        CellIsRule(operator="lessThan", formula=["0"],
                   fill=PatternFill("solid", fgColor=OVERBILLED)),
    )
    ws.conditional_formatting.add(
        span,
        CellIsRule(operator="greaterThan", formula=["0"],
                   fill=PatternFill("solid", fgColor=UNDERBILLED)),
    )


def _set_widths(ws):
    widths = {
        "A": 10, "B": 26, "C": 16, "D": 18, "E": 14, "F": 14,
        "G": 14, "H": 16, "I": 16, "J": 16, "K": 18, "L": 18, "M": 13,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def write_dashboard_sheet(ws, row_count):
    ws.title = "Dashboard"
    first = formulas.FIRST_DATA_ROW
    last = first + row_count - 1

    title = ws["A1"]
    title.value = "Construction WIP Dashboard"
    title.font = Font(bold=True, size=14, color="FFFFFF")
    title.fill = HEADER_FILL
    title.alignment = Alignment(vertical="center")
    ws.merge_cells("A1:B1")
    ws["B1"].fill = HEADER_FILL
    ws.row_dimensions[1].height = 24

    r = 3
    for label, formula, kind in formulas.dashboard_rows(first, last):
        ws["A%d" % r].value = label
        ws["A%d" % r].font = Font(bold=True)
        value_cell = ws["B%d" % r]
        value_cell.value = formula
        if kind == "money":
            value_cell.number_format = MONEY_FORMAT
        value_cell.alignment = Alignment(horizontal="right")
        r += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18


def build(schedule_path, out_path):
    rows = read_schedule(schedule_path)
    if not rows:
        raise ValueError("schedule file %s has no rows" % schedule_path)

    wb = Workbook()
    schedule = wb.active
    write_schedule_sheet(schedule, rows)
    dashboard = wb.create_sheet("Dashboard")
    write_dashboard_sheet(dashboard, len(rows))
    wb.active = wb.index(schedule)
    wb.save(out_path)
    return len(rows)


def main(argv=None):
    parser = argparse.ArgumentParser(description="Build the WIP workbook from the schedule CSV.")
    parser.add_argument("--schedule", default="wip_schedule.csv")
    parser.add_argument("--out", default="wip_workbook.xlsx")
    args = parser.parse_args(argv)

    count = build(args.schedule, args.out)
    print("Built %s from %s (%d jobs)." % (args.out, args.schedule, count))
    print("WIP Schedule sheet: inputs in columns C to F, live formulas in G to M.")
    print("Dashboard sheet: schedule totals and the job counts by billing position.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
