"""The layout and cell formulas for the chargeback workbook.

One source of truth shared by the builder and the verifier. The builder writes the
formulas into the Allocation sheet; the verifier regenerates the expected formula
for every cell and confirms the workbook holds exactly that, then computes it.

The Allocation sheet is laid out as:

  Row 1            header: Department, Driver, one column per cost item, Total
  Row 2 (POOL_ROW) pool amounts: the total driver in B, each item amount in its
                   column, the pool total in the Total column
  Rows 3..        one per department: the department, its driver value, a live
                   formula for its share of each item, and a row total
  Last row         column totals

A department's share of an item is item amount times the department's driver over
the total driver, rounded to the cent: =ROUND(item$2 * $Brow / $B$2, 2).
"""

from openpyxl.utils import get_column_letter

DEPT_COL = "A"
DRIVER_COL = "B"
FIRST_ITEM_COL_INDEX = 3   # column C
HEADER_ROW = 1
POOL_ROW = 2
FIRST_DEPT_ROW = 3


def item_columns(item_count):
    """Column letters that hold the item allocations, left to right."""
    return [get_column_letter(FIRST_ITEM_COL_INDEX + i) for i in range(item_count)]


def total_column(item_count):
    """Column letter for the per-row total, just past the last item."""
    return get_column_letter(FIRST_ITEM_COL_INDEX + item_count)


def allocation_formula(item_col, dept_row):
    """Live formula for one department's share of one item."""
    return "=ROUND({c}${p}*${d}{r}/${d}${p},2)".format(
        c=item_col, p=POOL_ROW, d=DRIVER_COL, r=dept_row)


def row_total_formula(item_cols, dept_row):
    """Sum of a department's item shares."""
    return "=SUM({a}{r}:{b}{r})".format(a=item_cols[0], b=item_cols[-1], r=dept_row)


def column_total_formula(col, first_row, last_row):
    """Sum of a column across the department rows."""
    return "=SUM({c}{a}:{c}{b})".format(c=col, a=first_row, b=last_row)


def dashboard_rows(total_col, totals_row, first_dept_row, last_dept_row, departments):
    """(label, formula, kind) rows for the Dashboard sheet, referencing Allocation."""
    sheet = "'Allocation'"
    rows = [
        ("Total pool", "=%s!%s%d" % (sheet, total_col, POOL_ROW), "money"),
        ("Total allocated", "=%s!%s%d" % (sheet, total_col, totals_row), "money"),
        ("Departments", len(departments), "count"),
    ]
    for i, dept in enumerate(departments):
        r = first_dept_row + i
        rows.append((dept + " chargeback", "=%s!%s%d" % (sheet, total_col, r), "money"))
    return rows
