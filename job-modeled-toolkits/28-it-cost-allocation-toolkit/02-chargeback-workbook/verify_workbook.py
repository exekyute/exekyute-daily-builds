"""Checks the chargeback workbook against the engine's allocation, to the cent.

The workbook holds live formulas, not baked-in numbers, so this tool proves two
things for every department and item. First, that each allocation cell holds the
formula it should. Second, that the formula, computed straight from the workbook's
own pool and driver cells, equals the engine's allocation to the cent. The
computation runs through the small evaluator in formula_eval.py, never through
Excel and never through the engine code, so it is an independent check.

It also confirms the totals row and the Dashboard sheet add up to the pool.

Usage:
  python verify_workbook.py
  python verify_workbook.py --matrix allocation_matrix.csv --workbook chargeback_workbook.xlsx
"""

import argparse
import csv
from decimal import Decimal, ROUND_HALF_UP

from openpyxl import load_workbook

import formulas
from formula_eval import Evaluator

CENT = Decimal("0.01")


def money(value):
    return Decimal(value).quantize(CENT, rounding=ROUND_HALF_UP)


def read_matrix(path):
    with open(path, newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        rows = list(reader)
        fields = reader.fieldnames
    items = [f for f in fields if f not in ("department", "driver_value", "total")]
    return rows, items


class Report:
    def __init__(self):
        self.passed = 0
        self.failures = []

    def check(self, ok, message):
        if ok:
            self.passed += 1
        else:
            self.failures.append(message)


def verify(matrix_path, workbook_path):
    rows, items = read_matrix(matrix_path)
    wb = load_workbook(workbook_path)
    ws = wb["Allocation"]
    ev = Evaluator(wb, "Allocation")
    report = Report()

    item_cols = formulas.item_columns(len(items))
    total_col = formulas.total_column(len(items))
    first = formulas.FIRST_DEPT_ROW
    last = first + len(rows) - 1
    totals_row = last + 1

    print("Verifying %s against %s" % (workbook_path, matrix_path))
    print()

    # Pool amount row: each item amount is the engine's column total.
    for item, col in zip(items, item_cols):
        column_total = money(sum(Decimal(r[item]) for r in rows))
        cell = ws["%s%d" % (col, formulas.POOL_ROW)]
        report.check(money(Decimal(str(cell.value))) == column_total,
                     "pool amount for %s is %s, expected %s" % (item, cell.value, column_total))

    pool_total = money(sum(Decimal(r["total"]) for r in rows))

    # Department rows.
    for i, row in enumerate(rows):
        r = first + i
        dept = row["department"]
        report.check(ws["%s%d" % (formulas.DEPT_COL, r)].value == dept, "%s: name mismatch" % dept)
        report.check(money(Decimal(str(ws["%s%d" % (formulas.DRIVER_COL, r)].value))) == money(row["driver_value"]),
                     "%s: driver mismatch" % dept)
        for item, col in zip(items, item_cols):
            cell = ws["%s%d" % (col, r)]
            expected = formulas.allocation_formula(col, r)
            report.check(cell.value == expected,
                         "%s/%s: formula is %r, expected %r" % (dept, item, cell.value, expected))
            computed = ev.evaluate(cell.value)
            report.check(money(computed) == money(row[item]),
                         "%s/%s: value %s, expected %s" % (dept, item, money(computed), money(row[item])))
        tcell = ws["%s%d" % (total_col, r)]
        report.check(money(ev.evaluate(tcell.value)) == money(row["total"]),
                     "%s: row total mismatch" % dept)
        print("  %-14s allocated %14s" % (dept, money(row["total"])))

    # Totals row.
    for item, col in zip(items, item_cols):
        column_total = money(sum(Decimal(r[item]) for r in rows))
        report.check(money(ev.evaluate(ws["%s%d" % (col, totals_row)].value)) == column_total,
                     "totals: column %s mismatch" % col)
    report.check(money(ev.evaluate(ws["%s%d" % (total_col, totals_row)].value)) == pool_total,
                 "totals: pool total mismatch")
    print()
    print("  Pool total %s, allocated %s" % (pool_total, pool_total))

    # Dashboard.
    _verify_dashboard(wb, rows, items, pool_total, report)

    print()
    if report.failures:
        print("FAIL: %d check(s) failed." % len(report.failures))
        for message in report.failures:
            print("  - %s" % message)
        return 1
    print("PASS: %d checks. Every live formula reproduces the engine to the cent." % report.passed)
    return 0


def _verify_dashboard(wb, rows, items, pool_total, report):
    ws = wb["Dashboard"]
    ev = Evaluator(wb, "Dashboard")
    by_label = {}
    for r in range(3, 3 + 3 + len(rows)):
        label = ws["A%d" % r].value
        if label is None:
            continue
        by_label[label] = ws["B%d" % r].value

    report.check(money(ev.evaluate(by_label["Total pool"])) == pool_total, "dashboard: total pool")
    report.check(money(ev.evaluate(by_label["Total allocated"])) == pool_total, "dashboard: total allocated")
    report.check(int(by_label["Departments"]) == len(rows), "dashboard: department count")
    for row in rows:
        label = row["department"] + " chargeback"
        report.check(money(ev.evaluate(by_label[label])) == money(row["total"]),
                     "dashboard: %s mismatch" % label)


def main(argv=None):
    parser = argparse.ArgumentParser(description="Verify the chargeback workbook against the allocation matrix.")
    parser.add_argument("--matrix", default="allocation_matrix.csv")
    parser.add_argument("--workbook", default="chargeback_workbook.xlsx")
    args = parser.parse_args(argv)
    return verify(args.matrix, args.workbook)


if __name__ == "__main__":
    raise SystemExit(main())
