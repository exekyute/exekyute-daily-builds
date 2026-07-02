"""Builds the chargeback workbook from the engine's allocation matrix.

Reads allocation_matrix.csv (written by the allocation engine in 01) and writes
chargeback_workbook.xlsx: an Allocation sheet whose department driver values are
written as numbers and whose allocation cells are live Excel formulas, plus a
Dashboard sheet that totals the pool and lists each department's chargeback.
Because the allocation cells are formulas, changing a driver recomputes the split.

Usage:
  python build_workbook.py
  python build_workbook.py --matrix allocation_matrix.csv --out chargeback_workbook.xlsx

Requires openpyxl (see requirements.txt). The engine in 01 is standard-library only.
"""

import argparse
import csv
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

import formulas

BASE = "1F3A5F"
ACCENT = "DCE6F1"
MONEY_FORMAT = '$#,##0.00'

HEADER_FONT = Font(bold=True, color="FFFFFF")
TOTAL_FONT = Font(bold=True)
HEADER_FILL = PatternFill("solid", fgColor=BASE)
TOTAL_FILL = PatternFill("solid", fgColor=ACCENT)


def read_matrix(path):
    with open(path, newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        rows = list(reader)
        fields = reader.fieldnames
    items = [f for f in fields if f not in ("department", "driver_value", "total")]
    departments = [r["department"] for r in rows]
    total_driver = sum(Decimal(r["driver_value"]) for r in rows)
    item_amounts = {
        item: sum(Decimal(r[item]) for r in rows) for item in items
    }
    drivers = {r["department"]: Decimal(r["driver_value"]) for r in rows}
    return {
        "rows": rows, "items": items, "departments": departments,
        "total_driver": total_driver, "item_amounts": item_amounts, "drivers": drivers,
    }


def write_allocation_sheet(ws, data):
    ws.title = "Allocation"
    items = data["items"]
    item_cols = formulas.item_columns(len(items))
    total_col = formulas.total_column(len(items))
    first = formulas.FIRST_DEPT_ROW
    last = first + len(data["departments"]) - 1
    totals_row = last + 1

    # Header row.
    headers = ["Department", "Driver"] + items + ["Total"]
    columns = [formulas.DEPT_COL, formulas.DRIVER_COL] + item_cols + [total_col]
    for col, text in zip(columns, headers):
        cell = ws["%s%d" % (col, formulas.HEADER_ROW)]
        cell.value = text
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Pool amount row.
    pool = formulas.POOL_ROW
    ws["%s%d" % (formulas.DEPT_COL, pool)].value = "Pool amount"
    ws["%s%d" % (formulas.DEPT_COL, pool)].font = TOTAL_FONT
    ws["%s%d" % (formulas.DRIVER_COL, pool)].value = float(data["total_driver"])
    for item, col in zip(items, item_cols):
        cell = ws["%s%d" % (col, pool)]
        cell.value = float(data["item_amounts"][item])
        cell.number_format = MONEY_FORMAT
        cell.font = TOTAL_FONT
    total_pool = ws["%s%d" % (total_col, pool)]
    total_pool.value = "=SUM(%s%d:%s%d)" % (item_cols[0], pool, item_cols[-1], pool)
    total_pool.number_format = MONEY_FORMAT
    total_pool.font = TOTAL_FONT

    # Department rows.
    for i, row in enumerate(data["rows"]):
        r = first + i
        ws["%s%d" % (formulas.DEPT_COL, r)].value = row["department"]
        ws["%s%d" % (formulas.DRIVER_COL, r)].value = float(Decimal(row["driver_value"]))
        for item, col in zip(items, item_cols):
            cell = ws["%s%d" % (col, r)]
            cell.value = formulas.allocation_formula(col, r)
            cell.number_format = MONEY_FORMAT
        tcell = ws["%s%d" % (total_col, r)]
        tcell.value = formulas.row_total_formula(item_cols, r)
        tcell.number_format = MONEY_FORMAT

    # Totals row.
    ws["%s%d" % (formulas.DEPT_COL, totals_row)].value = "Total"
    ws["%s%d" % (formulas.DEPT_COL, totals_row)].font = TOTAL_FONT
    for col in [formulas.DRIVER_COL] + item_cols + [total_col]:
        cell = ws["%s%d" % (col, totals_row)]
        cell.value = formulas.column_total_formula(col, first, last)
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        if col != formulas.DRIVER_COL:
            cell.number_format = MONEY_FORMAT
    ws["%s%d" % (formulas.DEPT_COL, totals_row)].fill = TOTAL_FILL

    _set_widths(ws, columns)
    return {"total_col": total_col, "totals_row": totals_row, "first": first, "last": last}


def _set_widths(ws, columns):
    ws.column_dimensions[columns[0]].width = 16
    ws.column_dimensions[columns[1]].width = 10
    for col in columns[2:]:
        ws.column_dimensions[col].width = 18


def write_dashboard_sheet(ws, data, layout):
    ws.title = "Dashboard"
    title = ws["A1"]
    title.value = "IT Showback Dashboard"
    title.font = Font(bold=True, size=14, color="FFFFFF")
    title.fill = HEADER_FILL
    title.alignment = Alignment(vertical="center")
    ws.merge_cells("A1:B1")
    ws["B1"].fill = HEADER_FILL
    ws.row_dimensions[1].height = 24

    rows = formulas.dashboard_rows(
        layout["total_col"], layout["totals_row"], layout["first"], layout["last"], data["departments"])
    r = 3
    for label, value, kind in rows:
        ws["A%d" % r].value = label
        ws["A%d" % r].font = Font(bold=True)
        cell = ws["B%d" % r]
        cell.value = value
        if kind == "money":
            cell.number_format = MONEY_FORMAT
        cell.alignment = Alignment(horizontal="right")
        r += 1
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18


def build(matrix_path, out_path):
    data = read_matrix(matrix_path)
    wb = Workbook()
    layout = write_allocation_sheet(wb.active, data)
    write_dashboard_sheet(wb.create_sheet("Dashboard"), data, layout)
    wb.active = 0
    wb.save(out_path)
    return data


def main(argv=None):
    parser = argparse.ArgumentParser(description="Build the chargeback workbook from the allocation matrix.")
    parser.add_argument("--matrix", default="allocation_matrix.csv")
    parser.add_argument("--out", default="chargeback_workbook.xlsx")
    args = parser.parse_args(argv)
    data = build(args.matrix, args.out)
    print("Built %s from %s (%d departments, %d items)." % (
        args.out, args.matrix, len(data["departments"]), len(data["items"])))
    print("Allocation sheet: drivers as values, allocation cells as live formulas.")
    print("Dashboard sheet: pool total and each department's chargeback.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
