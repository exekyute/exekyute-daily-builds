"""Thin driver for the Crime snapshot workbook.

This file holds no analytical shortcut. It reads the pinned CSV snapshot in
data/raw/, writes an Excel workbook whose every summary figure is a live cell
formula (COUNTIF / SUMPRODUCT / INDEX-MATCH against the data sheet, no pasted
results), and recomputes the same key figures independently in plain Python to
form the golden. The workbook is the deliverable; the golden is the check.

Usage:
    python build.py            regenerate the .xlsx from the snapshot, then verify
    python build.py verify      re-run only the golden diff (recompute vs expected/)
    python build.py show        print the key-figures table in the terminal

The golden lives in expected/key_figures.csv. It is recomputed in plain Python,
never read back from the workbook. Any share is rounded half-away-from-zero with
decimal.ROUND_HALF_UP to mirror Excel's ROUND, never Python's built-in round.
"""

import csv
import os
import sys
import zipfile
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

HERE = os.path.dirname(os.path.abspath(__file__))
RAW_PATH = os.path.join(HERE, "data", "raw", "hrm_crime_2026-07-13.csv")
XLSX_PATH = os.path.join(HERE, "crime_snapshot.xlsx")
OUT_DIR = os.path.join(HERE, "out")
OUT_KF_PATH = os.path.join(OUT_DIR, "key_figures.csv")
EXPECTED_KF_PATH = os.path.join(HERE, "expected", "key_figures.csv")

PULL_DATE = "2026-07-13"

# Fixed workbook metadata so a regenerated file is byte-stable across runs.
WORKBOOK_STAMP = datetime(2026, 7, 13)
# Fixed zip-entry timestamp: Python's zipfile stamps each entry with the current
# time by default, so the .xlsx bytes would drift between runs without this.
ZIP_STAMP = (2026, 7, 13, 0, 0, 0)

KF_HEADER = ["figure", "category", "count", "share_pct"]


# --------------------------------------------------------------------------
# Snapshot load
# --------------------------------------------------------------------------

def load_snapshot():
    """Read the pinned CSV and return incident rows sorted by the stable key
    (evt_date, evt_rin). Pure data preparation: parse types, keep the columns
    the workbook uses, and order deterministically. No aggregation here."""
    rows = []
    with open(RAW_PATH, "r", encoding="utf-8", newline="") as handle:
        for rec in csv.DictReader(handle):
            evt_date = datetime.strptime(
                rec["EVT_DATE"].strip(), "%m/%d/%Y %I:%M:%S %p").date()
            rows.append({
                "evt_date": evt_date,
                "evt_rin": int(rec["EVT_RIN"]),
                "category": rec["RUCR_EXT_D"].strip(),
                "code": int(rec["RUCR"]),
                "location": rec["LOCATION"].strip(),
            })
    rows.sort(key=lambda r: (r["evt_date"], r["evt_rin"]))
    return rows


def distinct_sorted(values):
    """Distinct values in a deterministic alphabetical order. Getting the axis
    labels is data preparation, not analysis: no counting happens here."""
    return sorted(set(values))


def normalize_zip_timestamps(path):
    """Rewrite the .xlsx so every zip entry carries a fixed timestamp, preserving
    entry order and content. Makes a regenerated workbook byte-identical."""
    tmp = path + ".tmp"
    with zipfile.ZipFile(path, "r") as zin:
        with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
            for info in zin.infolist():
                data = zin.read(info.filename)
                fresh = zipfile.ZipInfo(info.filename, date_time=ZIP_STAMP)
                fresh.compress_type = info.compress_type
                fresh.external_attr = info.external_attr
                fresh.internal_attr = info.internal_attr
                fresh.create_system = info.create_system
                zout.writestr(fresh, data)
    os.replace(tmp, path)


# --------------------------------------------------------------------------
# Golden key figures, recomputed in plain Python (never read from the workbook)
# --------------------------------------------------------------------------

def share_pct(count, total):
    """count / total as a percent, one decimal, half-away-from-zero. Mirrors the
    Excel ROUND(100*count/total, 1) written into the workbook."""
    if total == 0:
        return Decimal("0.0")
    return (Decimal(count) * Decimal(100) / Decimal(total)).quantize(
        Decimal("0.1"), rounding=ROUND_HALF_UP)


def compute_key_figures(rows):
    """Recompute every key figure the workbook exposes, in plain Python.

    Returns a list of [figure, category, count, share_pct] records in a fixed
    order: the total, the top category, then one row per category alphabetically
    (matching the workbook's summary rows)."""
    total = len(rows)
    categories = distinct_sorted(r["category"] for r in rows)
    counts = {c: sum(1 for r in rows if r["category"] == c) for c in categories}

    # Top category: highest count, ties broken by category name (alphabetical),
    # so the result is deterministic on any snapshot.
    top = min(categories, key=lambda c: (-counts[c], c))

    records = []
    records.append(["total", "", total, str(share_pct(total, total))])
    records.append(["top_category", top, counts[top], str(share_pct(counts[top], total))])
    for c in categories:
        records.append(["category", c, counts[c], str(share_pct(counts[c], total))])
    return records


def kf_lines(records):
    """Render the key-figure records as CSV text lines (header + rows)."""
    lines = [",".join(KF_HEADER)]
    for figure, category, count, share in records:
        lines.append(",".join([figure, category, str(count), share]))
    return lines


# --------------------------------------------------------------------------
# Workbook generation: data and live formulas only
# --------------------------------------------------------------------------

def generate_workbook(rows):
    """Write crime_snapshot.xlsx. The data sheet holds the raw incidents; the
    summary sheet holds only labels and live formulas that reference the data
    sheet. No analytical result is written as a value."""
    first_row = 2
    last_row = first_row + len(rows) - 1  # data occupies rows 2..(1+N)
    cat_rng = "data!$C${0}:$C${1}".format(first_row, last_row)
    loc_rng = "data!$E${0}:$E${1}".format(first_row, last_row)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F3B57")
    label_font = Font(bold=True)
    title_font = Font(bold=True, size=14)
    thin = Side(style="thin", color="B7C4D0")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    right = Alignment(horizontal="right")

    wb = Workbook()
    wb.properties.created = WORKBOOK_STAMP
    wb.properties.modified = WORKBOOK_STAMP

    # ---- data sheet -------------------------------------------------------
    data = wb.active
    data.title = "data"
    data_headers = ["evt_date", "evt_rin", "category", "code", "location"]
    for col, name in enumerate(data_headers, start=1):
        cell = data.cell(row=1, column=col, value=name)
        cell.font = header_font
        cell.fill = header_fill
    for i, r in enumerate(rows, start=first_row):
        d = data.cell(row=i, column=1, value=r["evt_date"])
        d.number_format = "yyyy-mm-dd"
        data.cell(row=i, column=2, value=r["evt_rin"])
        data.cell(row=i, column=3, value=r["category"])
        data.cell(row=i, column=4, value=r["code"])
        data.cell(row=i, column=5, value=r["location"])
    for col, width in zip("ABCDE", (13, 11, 22, 8, 26)):
        data.column_dimensions[col].width = width
    data.freeze_panes = "A2"

    # ---- summary sheet ----------------------------------------------------
    s = wb.create_sheet("summary")
    s["A1"] = "Halifax crime snapshot"
    s["A1"].font = title_font
    s["A2"] = "Snapshot pulled {0}; incidents dated 2026-07-05 to 2026-07-11.".format(PULL_DATE)
    s["A3"] = "Every count and share below is a live formula against the data sheet."

    # Headline: total incidents (independent count of the data sheet).
    s["A5"] = "Total incidents"
    s["A5"].font = label_font
    s["B5"] = "=COUNTA({0})".format(cat_rng)
    s["B5"].alignment = right

    # By category (alphabetical axis; counts and shares via formula).
    categories = distinct_sorted(r["category"] for r in rows)
    cat_head = 7
    s.cell(row=cat_head, column=1, value="Category").font = header_font
    s.cell(row=cat_head, column=2, value="Incidents").font = header_font
    s.cell(row=cat_head, column=3, value="Share %").font = header_font
    for col in (1, 2, 3):
        s.cell(row=cat_head, column=col).fill = header_fill
    first_cat = cat_head + 1
    for offset, cat in enumerate(categories):
        r = first_cat + offset
        s.cell(row=r, column=1, value=cat).border = box
        cnt = s.cell(row=r, column=2, value="=COUNTIF({0},A{1})".format(cat_rng, r))
        cnt.alignment = right
        cnt.border = box
        shr = s.cell(row=r, column=3, value="=ROUND(100*B{0}/$B$5,1)".format(r))
        shr.alignment = right
        shr.border = box
    last_cat = first_cat + len(categories) - 1
    total_row = last_cat + 1
    s.cell(row=total_row, column=1, value="Total").font = label_font
    tcnt = s.cell(row=total_row, column=2,
                  value="=SUM(B{0}:B{1})".format(first_cat, last_cat))
    tcnt.font = label_font
    tcnt.alignment = right
    tshr = s.cell(row=total_row, column=3,
                  value="=ROUND(100*B{0}/$B$5,1)".format(total_row))
    tshr.font = label_font
    tshr.alignment = right

    # Top category (found by formula, not by row position).
    top_lbl = total_row + 2
    s.cell(row=top_lbl, column=1, value="Top category").font = label_font
    s.cell(row=top_lbl, column=2,
           value="=INDEX($A${0}:$A${1},MATCH(MAX($B${0}:$B${1}),$B${0}:$B${1},0))".format(
               first_cat, last_cat)).alignment = right
    s.cell(row=top_lbl + 1, column=1, value="Top category incidents").font = label_font
    s.cell(row=top_lbl + 1, column=2,
           value="=MAX($B${0}:$B${1})".format(first_cat, last_cat)).alignment = right
    s.cell(row=top_lbl + 2, column=1, value="Top category share %").font = label_font
    s.cell(row=top_lbl + 2, column=2,
           value="=ROUND(100*B{0}/$B$5,1)".format(top_lbl + 1)).alignment = right

    for col, width in zip("ABC", (24, 11, 9)):
        s.column_dimensions[col].width = width

    # By area (location), alphabetical axis; counts via formula.
    area = wb.create_sheet("by_area")
    area["A1"] = "Incidents by area (location)"
    area["A1"].font = title_font
    ah = 3
    area.cell(row=ah, column=1, value="Location").font = header_font
    area.cell(row=ah, column=2, value="Incidents").font = header_font
    for col in (1, 2):
        area.cell(row=ah, column=col).fill = header_fill
    locations = distinct_sorted(r["location"] for r in rows)
    first_loc = ah + 1
    for offset, loc in enumerate(locations):
        r = first_loc + offset
        area.cell(row=r, column=1, value=loc)
        area.cell(row=r, column=2,
                  value="=COUNTIF({0},A{1})".format(loc_rng, r)).alignment = right
    last_loc = first_loc + len(locations) - 1
    lt = last_loc + 1
    area.cell(row=lt, column=1, value="Total").font = label_font
    area.cell(row=lt, column=2,
              value="=SUM(B{0}:B{1})".format(first_loc, last_loc)).font = label_font
    area.column_dimensions["A"].width = 26
    area.column_dimensions["B"].width = 11

    wb.save(XLSX_PATH)
    normalize_zip_timestamps(XLSX_PATH)


# --------------------------------------------------------------------------
# verify / run / show
# --------------------------------------------------------------------------

def read_lines(path):
    with open(path, "r", encoding="utf-8", newline="") as handle:
        return handle.read().splitlines()


def verify():
    """Recompute the key figures from the snapshot in plain Python and diff them
    against expected/key_figures.csv. PASS only on an exact match."""
    if not os.path.exists(EXPECTED_KF_PATH):
        print("FAIL: expected/key_figures.csv does not exist.")
        return False
    actual = kf_lines(compute_key_figures(load_snapshot()))
    expected = read_lines(EXPECTED_KF_PATH)
    if actual == expected:
        print("PASS: recomputed key figures match expected/key_figures.csv ({0} figures).".format(
            max(len(actual) - 1, 0)))
        return True
    print("FAIL: recomputed key figures differ from expected/key_figures.csv.")
    print("  expected {0} lines, got {1} lines.".format(len(expected), len(actual)))
    limit = max(len(actual), len(expected))
    shown = 0
    for i in range(limit):
        a = actual[i] if i < len(actual) else "<missing>"
        e = expected[i] if i < len(expected) else "<missing>"
        if a != e:
            print("  line {0}:".format(i + 1))
            print("    expected: {0}".format(e))
            print("    actual:   {0}".format(a))
            shown += 1
            if shown >= 8:
                print("  ... (further differences suppressed)")
                break
    return False


def run():
    """Regenerate the workbook from the snapshot, write the recomputed golden to
    out/, then verify it against expected/."""
    rows = load_snapshot()
    generate_workbook(rows)
    print("wrote {0} ({1} incident rows)".format(os.path.basename(XLSX_PATH), len(rows)))
    os.makedirs(OUT_DIR, exist_ok=True)
    lines = kf_lines(compute_key_figures(rows))
    with open(OUT_KF_PATH, "w", encoding="utf-8", newline="") as handle:
        handle.write("\n".join(lines) + "\n")
    print("wrote out/key_figures.csv")
    print("")
    return verify()


def print_table(columns, rows, numeric_cols):
    """Aligned plain-ASCII table. cp1252 safe: no box-drawing characters.
    Columns flagged numeric are right-aligned."""
    cells = [[("" if v is None else str(v)) for v in row] for row in rows]
    widths = [len(columns[i]) for i in range(len(columns))]
    for row in cells:
        for i, text in enumerate(row):
            widths[i] = max(widths[i], len(text))

    def line(values):
        parts = []
        for i, text in enumerate(values):
            parts.append(text.rjust(widths[i]) if i in numeric_cols else text.ljust(widths[i]))
        return "  ".join(parts)

    print(line(columns))
    print("  ".join("-" * w for w in widths))
    for row in cells:
        print(line(row))


def show():
    """Print the key-figures table. Recomputes from the snapshot (the same figures
    the workbook exposes as formulas); falls back to expected/ if no snapshot."""
    if os.path.exists(RAW_PATH):
        records = compute_key_figures(load_snapshot())
    elif os.path.exists(EXPECTED_KF_PATH):
        records = [ln.split(",") for ln in read_lines(EXPECTED_KF_PATH)[1:]]
        records = [[r[0], r[1], int(r[2]), r[3]] for r in records]
    else:
        print("Nothing to show yet. Run: python build.py")
        return

    total = next(r for r in records if r[0] == "total")
    top = next(r for r in records if r[0] == "top_category")
    cats = [r for r in records if r[0] == "category"]

    print("\nCrime snapshot key figures (snapshot {0}, incidents 2026-07-05 to 2026-07-11)\n".format(
        PULL_DATE))
    print("Total incidents: {0}".format(total[2]))
    print("Top category:    {0} ({1}, {2}%)\n".format(top[1], top[2], top[3]))

    table_rows = [[c[1], c[2], c[3]] for c in cats]
    table_rows.append(["Total", total[2], total[3]])
    print_table(["Category", "Incidents", "Share %"], table_rows, numeric_cols={1, 2})
    print("")


def main():
    args = sys.argv[1:]
    if args and args[0] == "verify":
        sys.exit(0 if verify() else 1)
    if args and args[0] == "show":
        show()
        sys.exit(0)
    if args:
        print("unknown argument: {0}. Use no argument, 'verify', or 'show'.".format(args[0]))
        sys.exit(2)
    sys.exit(0 if run() else 1)


if __name__ == "__main__":
    main()
