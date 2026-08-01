"""Thin driver for the HRM area rate cost calculator workbook.

This file holds no analytical shortcut. It reads two pinned CSV snapshots in
data/raw/ (the Tax Rates table and the enumerated area-rate polygons), writes an
Excel workbook whose every charge is a live cell formula (INDEX/MATCH against a
rates sheet, no pasted results), and recomputes the same worked examples
independently in plain Python to form the golden. The workbook is the deliverable;
the golden is the check.

Usage:
    python build.py            regenerate the .xlsx from the snapshots, then verify
    python build.py verify      re-run only the golden diff (recompute vs expected/)
    python build.py show        print the worked-example figures in the terminal

The golden lives in expected/key_figures.csv. It is recomputed in plain Python,
never read back from the workbook. Money is rounded half-away-from-zero with
decimal.ROUND_HALF_UP to mirror Excel's ROUND, never Python's built-in round.

The join: an area's AREARATE_CODE equals a Tax_Rates Rate_Code (verified live:
area A000 Frame Subdivision Homeowners maps to Rate_Code A000). A charge is looked
up on Rate_Code plus property class (Rate_Type). Rate basis: a Calculation_Type of
"Rate" is dollars per $100 of taxable assessment; "Flat Rate" is a fixed dollar
charge per property.
"""

import csv
import os
import re
import sys
import zipfile
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

HERE = os.path.dirname(os.path.abspath(__file__))
RATES_RAW = os.path.join(HERE, "data", "raw", "hrm_tax-rates_2026-07-13.csv")
AREAS_RAW = os.path.join(HERE, "data", "raw", "hrm_area-rates_2026-07-13.csv")
XLSX_PATH = os.path.join(HERE, "area_rate_calculator.xlsx")
OUT_DIR = os.path.join(HERE, "out")
OUT_KF_PATH = os.path.join(OUT_DIR, "key_figures.csv")
EXPECTED_KF_PATH = os.path.join(HERE, "expected", "key_figures.csv")

PULL_DATE = "2026-07-13"
BILL_YEAR = 2025

# Fixed workbook metadata so a regenerated file is byte-stable across runs.
WORKBOOK_STAMP = datetime(2026, 7, 13)
# Fixed zip-entry timestamp: Python's zipfile stamps each entry with the current
# time by default, so the .xlsx bytes would drift between runs without this.
ZIP_STAMP = (2026, 7, 13, 0, 0, 0)

CLASSES = ["Residential", "Commercial", "Resource"]

# The documented worked examples. Each is one calculator scenario: a stated
# assessment, a property class, and the selected area codes stacked into a total.
# Scenario A stacks a flat community rate onto two per-$100 municipal rates;
# Scenario B switches the class to Commercial so the same lookup returns the
# Commercial rate. Both are recomputed in plain Python for the golden and shown
# live in the workbook.
SCENARIOS = [
    {"key": "A", "label": "Scenario A: Residential",
     "assessment": 325000, "klass": "Residential",
     "codes": ["M050", "M060", "A000"]},
    {"key": "B", "label": "Scenario B: Commercial",
     "assessment": 600000, "klass": "Commercial",
     "codes": ["M050", "R000", "A020"]},
]

KF_HEADER = ["scenario", "figure", "area_code", "class", "calc_type",
             "rate", "assessment", "charge", "share_pct"]


# --------------------------------------------------------------------------
# Snapshot load (pure data preparation: parse, filter, dedup, sort; no analysis)
# --------------------------------------------------------------------------

def ascii_clean(text):
    """Fold the few non-ASCII source glyphs (a curly apostrophe in one area
    name) to plain ASCII so the workbook and the terminal table stay cp1252
    safe. Data preparation, not analysis."""
    swaps = {"’": "'", "‘": "'", "“": '"', "”": '"'}
    for bad, good in swaps.items():
        text = text.replace(bad, good)
    return text.encode("ascii", "ignore").decode("ascii")


def load_rates():
    """Read the Tax Rates snapshot, keep the latest bill year, and return one
    record per (Rate_Code, Rate_Type), sorted by code then type. The Rate is kept
    as its exact source string so arithmetic uses Decimal, never a float."""
    rows = []
    with open(RATES_RAW, "r", encoding="utf-8", newline="") as handle:
        for rec in csv.DictReader(handle):
            if int(rec["Bill_Year"]) != BILL_YEAR:
                continue
            rows.append({
                "code": rec["Rate_Code"].strip(),
                "klass": rec["Rate_Type"].strip(),
                "rate": rec["Rate"].strip(),
                "calc": rec["Calculation_Type"].strip(),
                "desc": ascii_clean(rec["Rate_Description"].strip()),
            })
    rows.sort(key=lambda r: (r["code"], r["klass"]))
    return rows


def load_areas():
    """Read the area-rate snapshot and collapse it to one row per area code.

    A code can appear on several polygons (Transit M060 carries a label per year;
    one private road has two disjoint polygons). Getting the axis is preparation,
    not analysis: dedup to the distinct code, keep the last-sorting label so the
    result is deterministic, and record the layer category."""
    by_code = {}
    with open(AREAS_RAW, "r", encoding="utf-8", newline="") as handle:
        for rec in csv.DictReader(handle):
            code = rec["AREARATE_CODE"].strip()
            desc = ascii_clean(rec["DESCRIP"].strip())
            category = rec["category"].strip()
            if code not in by_code:
                by_code[code] = {"code": code, "desc": desc, "category": category}
            elif desc > by_code[code]["desc"]:
                by_code[code]["desc"] = desc
    return [by_code[c] for c in sorted(by_code)]


def rate_index(rates):
    """Map (code, class) -> record for the charge lookup."""
    return {(r["code"], r["klass"]): r for r in rates}


def normalize_zip_timestamps(path):
    """Rewrite the .xlsx so every zip entry carries a fixed timestamp and the core
    properties 'modified' date is pinned. openpyxl stamps 'modified' with the wall
    clock at save time regardless of the value set on the workbook, so without this
    the bytes drift between runs. Makes a regenerated workbook byte-identical."""
    stamp = WORKBOOK_STAMP.strftime("%Y-%m-%dT%H:%M:%SZ")
    tmp = path + ".tmp"
    with zipfile.ZipFile(path, "r") as zin:
        with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
            for info in zin.infolist():
                data = zin.read(info.filename)
                if info.filename == "docProps/core.xml":
                    text = data.decode("utf-8")
                    text = re.sub(
                        r"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)",
                        r"\g<1>" + stamp + r"\g<2>", text)
                    data = text.encode("utf-8")
                fresh = zipfile.ZipInfo(info.filename, date_time=ZIP_STAMP)
                fresh.compress_type = info.compress_type
                fresh.external_attr = info.external_attr
                fresh.internal_attr = info.internal_attr
                fresh.create_system = info.create_system
                zout.writestr(fresh, data)
    os.replace(tmp, path)


# --------------------------------------------------------------------------
# Golden worked examples, recomputed in plain Python (never read from the sheet)
# --------------------------------------------------------------------------

def money(value):
    """Round to the cent, half-away-from-zero, mirroring Excel ROUND(x, 2)."""
    return value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def share_pct(part, whole):
    """part / whole as a percent, one decimal, half-away-from-zero. Mirrors the
    Excel ROUND(100*part/whole, 1) written into the workbook."""
    if whole == 0:
        return Decimal("0.0")
    return (Decimal(100) * part / whole).quantize(
        Decimal("0.1"), rounding=ROUND_HALF_UP)


def charge_for(record, assessment):
    """The charge for one matched rate record. Flat Rate is the rate itself; Rate
    is dollars per $100 of assessment. Rounded to the cent."""
    rate = Decimal(record["rate"])
    if record["calc"] == "Flat Rate":
        raw = rate
    else:
        raw = rate * Decimal(assessment) / Decimal(100)
    return money(raw)


def compute_key_figures(rindex):
    """Recompute every worked-example figure the workbook exposes, in plain
    Python. Returns records [scenario, figure, area_code, class, calc_type, rate,
    assessment, charge, share_pct] in a fixed order: for each scenario, one row
    per selected code, then the scenario total."""
    records = []
    for sc in SCENARIOS:
        assessment = sc["assessment"]
        klass = sc["klass"]
        charges = []
        for code in sc["codes"]:
            rec = rindex[(code, klass)]
            charges.append((code, rec, charge_for(rec, assessment)))
        total = sum((c for _, _, c in charges), Decimal("0.00"))
        for code, rec, chg in charges:
            records.append([
                sc["key"], "charge", code, klass, rec["calc"], rec["rate"],
                str(assessment), str(money(chg)), str(share_pct(chg, total)),
            ])
        records.append([
            sc["key"], "total", "", klass, "", "", str(assessment),
            str(money(total)), str(share_pct(total, total)),
        ])
    return records


def kf_lines(records):
    """Render the key-figure records as CSV text lines (header + rows)."""
    lines = [",".join(KF_HEADER)]
    for rec in records:
        lines.append(",".join(str(v) for v in rec))
    return lines


# --------------------------------------------------------------------------
# Workbook generation: data and live formulas only
# --------------------------------------------------------------------------

def generate_workbook(rates, areas):
    """Write area_rate_calculator.xlsx. The rates and areas sheets hold the pinned
    reference data; the calculator sheet holds only inputs, labels, and live
    formulas that look up each charge and stack it into a total. No charge is
    written as a value."""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F3B57")
    label_font = Font(bold=True)
    title_font = Font(bold=True, size=14)
    band_fill = PatternFill("solid", fgColor="E8EEF4")
    thin = Side(style="thin", color="B7C4D0")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    right = Alignment(horizontal="right")

    wb = Workbook()
    wb.properties.created = WORKBOOK_STAMP
    wb.properties.modified = WORKBOOK_STAMP

    # ---- rates sheet ------------------------------------------------------
    # One row per (Rate_Code, Rate_Type) for the latest bill year. Column F is a
    # live composite key (code|class) the calculator matches against.
    rs = wb.active
    rs.title = "rates"
    rate_headers = ["Rate_Code", "Rate_Type", "Rate", "Calculation_Type",
                    "Rate_Description", "key"]
    for col, name in enumerate(rate_headers, start=1):
        c = rs.cell(row=1, column=col, value=name)
        c.font = header_font
        c.fill = header_fill
    r_first = 2
    for i, rec in enumerate(rates, start=r_first):
        rs.cell(row=i, column=1, value=rec["code"])
        rs.cell(row=i, column=2, value=rec["klass"])
        rc = rs.cell(row=i, column=3, value=float(rec["rate"]))
        rc.alignment = right
        rs.cell(row=i, column=4, value=rec["calc"])
        rs.cell(row=i, column=5, value=rec["desc"])
        rs.cell(row=i, column=6, value="=A{0}&\"|\"&B{0}".format(i))
    r_last = r_first + len(rates) - 1
    for col, width in zip("ABCDEF", (10, 13, 9, 17, 34, 20)):
        rs.column_dimensions[col].width = width
    rs.freeze_panes = "A2"
    key_rng = "rates!$F${0}:$F${1}".format(r_first, r_last)
    rate_col = "rates!$C${0}:$C${1}".format(r_first, r_last)
    calc_col = "rates!$D${0}:$D${1}".format(r_first, r_last)

    # ---- areas sheet ------------------------------------------------------
    # One row per area code, with its label and the layer category.
    ar = wb.create_sheet("areas")
    area_headers = ["AREARATE_CODE", "DESCRIP", "category"]
    for col, name in enumerate(area_headers, start=1):
        c = ar.cell(row=1, column=col, value=name)
        c.font = header_font
        c.fill = header_fill
    a_first = 2
    for i, rec in enumerate(areas, start=a_first):
        ar.cell(row=i, column=1, value=rec["code"])
        ar.cell(row=i, column=2, value=rec["desc"])
        ar.cell(row=i, column=3, value=rec["category"])
    a_last = a_first + len(areas) - 1
    for col, width in zip("ABC", (15, 48, 33)):
        ar.column_dimensions[col].width = width
    ar.freeze_panes = "A2"
    area_code_col = "areas!$A${0}:$A${1}".format(a_first, a_last)
    area_desc_col = "areas!$B${0}:$B${1}".format(a_first, a_last)

    # A defined name for the area-code list drives the code dropdowns.
    wb.defined_names["AreaCodes"] = DefinedName(
        "AreaCodes", attr_text="areas!$A${0}:$A${1}".format(a_first, a_last))

    # ---- calculator sheet (the face) --------------------------------------
    cs = wb.create_sheet("calculator")
    cs["A1"] = "HRM area rate cost calculator"
    cs["A1"].font = title_font
    cs["A2"] = ("Bill year {0}. Every charge below is a live formula against the "
                "rates sheet.".format(BILL_YEAR))
    cs["A3"] = ("Calculation_Type 'Rate' is dollars per $100 of taxable "
                "assessment; 'Flat Rate' is a fixed dollar charge per property.")

    col_widths = zip("ABCDEFGH", (11, 40, 20, 9, 12, 12, 9, 34))
    for col, width in col_widths:
        cs.column_dimensions[col].width = width

    class_dv = DataValidation(
        type="list", formula1='"{0}"'.format(",".join(CLASSES)), allow_blank=False)
    code_dv = DataValidation(type="list", formula1="AreaCodes", allow_blank=True)
    cs.add_data_validation(class_dv)
    cs.add_data_validation(code_dv)

    tbl_headers = ["Area code", "Description", "Lookup key", "Rate", "Calc type",
                   "Charge ($)", "Share %", "Note"]
    n_rows = 6  # code rows per scenario (seeded plus blanks the user can fill)

    cell_map = {}  # scenario key -> {"assessment","class","total","codes":{code:row}}
    row = 5
    for sc in SCENARIOS:
        cs.cell(row=row, column=1, value=sc["label"]).font = label_font
        title_row = row

        a_row = row + 1
        cs.cell(row=a_row, column=1, value="Assessment ($)").font = label_font
        acell = cs.cell(row=a_row, column=2, value=sc["assessment"])
        acell.number_format = "#,##0"
        acell.alignment = right
        assess_ref = "$B${0}".format(a_row)

        k_row = a_row + 1
        cs.cell(row=k_row, column=1, value="Property class").font = label_font
        kcell = cs.cell(row=k_row, column=2, value=sc["klass"])
        class_dv.add(kcell)
        class_ref = "$B${0}".format(k_row)

        head_row = k_row + 2
        for col, name in enumerate(tbl_headers, start=1):
            c = cs.cell(row=head_row, column=col, value=name)
            c.font = header_font
            c.fill = header_fill

        first = head_row + 1
        last = first + n_rows - 1
        total_row = last + 1
        total_ref = "$F${0}".format(total_row)
        codes = list(sc["codes"]) + [""] * (n_rows - len(sc["codes"]))
        code_rows = {}
        for offset, code in enumerate(codes):
            r = first + offset
            a = cs.cell(row=r, column=1, value=(code or None))
            a.border = box
            code_dv.add(a)
            # Description from the areas sheet.
            cs.cell(row=r, column=2,
                    value=('=IF($A{0}="","",IFERROR(INDEX({1},MATCH($A{0},{2},0)),'
                           '"unknown code"))'.format(r, area_desc_col, area_code_col))
                    ).border = box
            # Composite lookup key: code|class.
            cs.cell(row=r, column=3,
                    value='=IF($A{0}="","",$A{0}&"|"&{1})'.format(r, class_ref)
                    ).border = box
            # Rate and Calc type, matched on the composite key.
            rc = cs.cell(row=r, column=4,
                         value=('=IF($C{0}="","",IFERROR(INDEX({1},MATCH($C{0},{2},0)),'
                                '""))'.format(r, rate_col, key_rng)))
            rc.alignment = right
            rc.border = box
            ec = cs.cell(row=r, column=5,
                         value=('=IF($C{0}="","",IFERROR(INDEX({1},MATCH($C{0},{2},0)),'
                                '""))'.format(r, calc_col, key_rng)))
            ec.border = box
            # Charge: Flat Rate is the rate; otherwise rate * assessment / 100.
            fc = cs.cell(row=r, column=6,
                         value=('=IF($A{0}="","",IFERROR(ROUND(IF($E{0}="Flat Rate",'
                                '$D{0},$D{0}*{1}/100),2),0))'.format(r, assess_ref)))
            fc.number_format = "#,##0.00"
            fc.alignment = right
            fc.border = box
            # Share of the stacked total.
            gc = cs.cell(row=r, column=7,
                         value=('=IF($A{0}="","",IF({1}=0,0,ROUND(100*$F{0}/{1},1)))'
                                .format(r, total_ref)))
            gc.number_format = "0.0"
            gc.alignment = right
            gc.border = box
            # Note: flags a code/class with no rate for this bill year.
            cs.cell(row=r, column=8,
                    value=('=IF($A{0}="","",IF(ISNA(MATCH($C{0},{1},0)),'
                           '"no {2} rate for this code and class",""))'.format(
                               r, key_rng, BILL_YEAR))).border = box
            if code:
                code_rows[code] = r

        cs.cell(row=total_row, column=1, value="Total").font = label_font
        tf = cs.cell(row=total_row, column=6,
                     value="=SUM(F{0}:F{1})".format(first, last))
        tf.number_format = "#,##0.00"
        tf.font = label_font
        tf.alignment = right
        tg = cs.cell(row=total_row, column=7,
                     value="=IF({0}=0,0,ROUND(100*{0}/{0},1))".format(total_ref))
        tg.number_format = "0.0"
        tg.font = label_font
        tg.alignment = right

        for band in (title_row,):
            for col in range(1, 9):
                cs.cell(row=band, column=col).fill = band_fill

        cell_map[sc["key"]] = {
            "assessment": a_row, "class": k_row, "total": total_row,
            "codes": code_rows}
        row = total_row + 2

    cs["A1"].alignment = Alignment(vertical="center")

    wb.move_sheet("calculator", -(len(wb.sheetnames) - 1))  # calculator first
    wb.active = wb.sheetnames.index("calculator")

    wb.save(XLSX_PATH)
    normalize_zip_timestamps(XLSX_PATH)
    return cell_map


# --------------------------------------------------------------------------
# verify / run / show
# --------------------------------------------------------------------------

def read_lines(path):
    with open(path, "r", encoding="utf-8", newline="") as handle:
        return handle.read().splitlines()


def verify():
    """Recompute the worked examples from the snapshots in plain Python and diff
    them against expected/key_figures.csv. PASS only on an exact match."""
    if not os.path.exists(EXPECTED_KF_PATH):
        print("FAIL: expected/key_figures.csv does not exist.")
        return False
    actual = kf_lines(compute_key_figures(rate_index(load_rates())))
    expected = read_lines(EXPECTED_KF_PATH)
    if actual == expected:
        print("PASS: recomputed key figures match expected/key_figures.csv "
              "({0} figures).".format(max(len(actual) - 1, 0)))
        return True
    print("FAIL: recomputed key figures differ from expected/key_figures.csv.")
    print("  expected {0} lines, got {1} lines.".format(len(expected), len(actual)))
    limit = max(len(actual), len(expected))
    shown = 0
    for i in range(limit):
        a = actual[i] if i < len(actual) else "<missing>"
        e = expected[i] if i < len(expected) else "<missing>"
        if a != e:
            print("  line {0}:".format(i + 1))
            print("    expected: {0}".format(e))
            print("    actual:   {0}".format(a))
            shown += 1
            if shown >= 8:
                print("  ... (further differences suppressed)")
                break
    return False


def run():
    """Regenerate the workbook from the snapshots, write the recomputed golden to
    out/, then verify it against expected/."""
    rates = load_rates()
    areas = load_areas()
    generate_workbook(rates, areas)
    print("wrote {0} ({1} rate rows, {2} area codes)".format(
        os.path.basename(XLSX_PATH), len(rates), len(areas)))
    os.makedirs(OUT_DIR, exist_ok=True)
    lines = kf_lines(compute_key_figures(rate_index(rates)))
    with open(OUT_KF_PATH, "w", encoding="utf-8", newline="") as handle:
        handle.write("\n".join(lines) + "\n")
    print("wrote out/key_figures.csv")
    print("")
    return verify()


def print_table(columns, rows, numeric_cols):
    """Aligned plain-ASCII table. cp1252 safe: no box-drawing characters.
    Columns flagged numeric are right-aligned."""
    cells = [[("" if v is None else str(v)) for v in row] for row in rows]
    widths = [len(columns[i]) for i in range(len(columns))]
    for row in cells:
        for i, text in enumerate(row):
            widths[i] = max(widths[i], len(text))

    def line(values):
        parts = []
        for i, text in enumerate(values):
            parts.append(text.rjust(widths[i]) if i in numeric_cols else text.ljust(widths[i]))
        return "  ".join(parts)

    print(line(columns))
    print("  ".join("-" * w for w in widths))
    for row in cells:
        print(line(row))


def show():
    """Print the worked-example figures as aligned tables. Recomputes from the
    snapshots (the same figures the workbook exposes as formulas)."""
    if not os.path.exists(RATES_RAW):
        print("Nothing to show yet. Run: python build.py")
        return
    records = compute_key_figures(rate_index(load_rates()))
    by_scenario = {sc["key"]: sc for sc in SCENARIOS}

    print("\nHRM area rate cost calculator: worked examples (bill year {0})".format(
        BILL_YEAR))
    for sc in SCENARIOS:
        rows = [r for r in records if r[0] == sc["key"]]
        total = next(r for r in rows if r[1] == "total")
        charges = [r for r in rows if r[1] == "charge"]
        print("\n{0}  (assessment ${1:,}, class {2})".format(
            sc["label"], sc["assessment"], sc["klass"]))
        table = []
        for r in charges:
            table.append([r[2], r[4], r[5], r[7], r[8]])
        table.append(["Total", "", "", total[7], total[8]])
        print_table(
            ["Area code", "Calc type", "Rate", "Charge ($)", "Share %"],
            table, numeric_cols={2, 3, 4})
    print("")


def main():
    args = sys.argv[1:]
    if args and args[0] == "verify":
        sys.exit(0 if verify() else 1)
    if args and args[0] == "show":
        show()
        sys.exit(0)
    if args:
        print("unknown argument: {0}. Use no argument, 'verify', or 'show'.".format(args[0]))
        sys.exit(2)
    sys.exit(0 if run() else 1)


if __name__ == "__main__":
    main()
