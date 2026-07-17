"""Build and verify the public pay-band cost model workbook.

Commands:
    python build.py            rebuild the workbook, then verify the key figures
    python build.py verify     re-run the key-figure check only
    python build.py show       print the key figures as an aligned table

Reads the pinned snapshot in data/raw/, writes public-pay-band-cost-model.xlsx
with live Excel formulas (no VBA, no macros), then recomputes every key figure
in plain Python with decimal.Decimal and ROUND_HALF_UP (mirroring Excel's
ROUND) and diffs the result against expected/key_figures.csv.
"""

import csv
import sys
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

HERE = Path(__file__).resolve().parent
SNAPSHOT = HERE / "data" / "raw" / "ns_government-pay-scales_2026-07-06.csv"
EXPECTED = HERE / "expected" / "key_figures.csv"
WORKBOOK = HERE / "public-pay-band-cost-model.xlsx"

RAISE_PCT = Decimal("0.02")  # verification pins the Inputs cell at 2 percent
CENT = Decimal("0.01")


def money(value):
    """Round to the cent, half away from zero, matching Excel's ROUND."""
    return value.quantize(CENT, rounding=ROUND_HALF_UP)


def load_scale():
    """Read the snapshot and return the current pay-scale grid.

    Cleaning rules (documented in spec.md):
      1. Keep only rows from the latest scale period (max start_date).
      2. Drop rows with a blank biweekly_pay_rate.
      3. classification = pay_plan, step label = int(pay_plan_level),
         rate = Decimal(biweekly_pay_rate).
      4. (classification, step label) must be unique; rates must be
         cent-precise. Violations stop the build.
      5. Steps go into pay-progression order: labels 80 and up first
         (the below-range series), then the rest, each ascending. Along
         that progression rates must not fall; a falling rate stops the
         build.

    Returns an ordered dict: {classification: {step_label: rate}} sorted
    by classification, steps in progression order inside each
    classification.
    """
    if not SNAPSHOT.exists():
        sys.exit(f"snapshot not found: {SNAPSHOT.relative_to(HERE)}")

    with open(SNAPSHOT, newline="", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    if not rows:
        sys.exit("snapshot is empty")

    latest = max(r["start_date"] for r in rows)
    current = [r for r in rows if r["start_date"] == latest]

    grid = {}
    dropped_blank = 0
    for r in current:
        raw_rate = (r["biweekly_pay_rate"] or "").strip()
        if not raw_rate:
            dropped_blank += 1
            continue
        plan = r["pay_plan"].strip()
        step_text = r["pay_plan_level"].strip()
        if not step_text.isdigit():
            sys.exit(f"non-numeric pay_plan_level {step_text!r} for {plan}")
        step = int(step_text)
        rate = Decimal(raw_rate)
        if rate != money(rate):
            sys.exit(f"rate not cent-precise: {plan} step {step} = {raw_rate}")
        if plan in grid and step in grid[plan]:
            if grid[plan][step] != rate:
                sys.exit(f"conflicting rates for {plan} step {step}")
            continue
        grid.setdefault(plan, {})[step] = rate

    ordered = {}
    wrapped = 0
    for plan in sorted(grid):
        # Pay-progression order: 32 plans (EC, LM, SO) number their
        # below-range steps 80 to 99 and continue the main range from
        # label 0, so labels 80+ precede the rest.
        labels = sorted(grid[plan], key=lambda s: (s < 80, s))
        if labels != sorted(labels):
            wrapped += 1
        steps = {label: grid[plan][label] for label in labels}
        rates = list(steps.values())
        if any(b < a for a, b in zip(rates, rates[1:])):
            sys.exit(f"rates fall along the step progression for {plan}")
        ordered[plan] = steps

    meta = {
        "latest_period_start": latest[:10],
        "period_rows": len(current),
        "dropped_blank_rate": dropped_blank,
        "wrapped_plans": wrapped,
    }
    return ordered, meta


def classification_rows(grid):
    """Per-classification derived figures, mirroring the Model sheet row math."""
    out = []
    for plan, steps in grid.items():
        rates = list(steps.values())
        n = len(rates)
        first, top = min(rates), max(rates)
        span = top - first
        one_step = money(span / (n - 1)) if n > 1 else Decimal("0.00")
        base = sum(rates)
        raise_cost = sum(money(r * RAISE_PCT) for r in rates)
        out.append({
            "classification": plan,
            "steps": n,
            "first": first,
            "top": top,
            "span": span,
            "one_step": one_step,
            "base": base,
            "raise_cost": raise_cost,
        })
    return out


def compute_key_figures(grid):
    """The key figures, in the same order as the Model sheet cells."""
    rows = classification_rows(grid)
    total_base = sum(r["base"] for r in rows)
    raise_total = sum(r["raise_cost"] for r in rows)
    top_raise = max(rows, key=lambda r: r["raise_cost"])
    widest = max(rows, key=lambda r: r["span"])
    return [
        ("raise_pct", str(RAISE_PCT)),
        ("classification_count", str(len(rows))),
        ("published_rate_count", str(sum(r["steps"] for r in rows))),
        ("total_biweekly_base", str(money(total_base))),
        ("raise_total", str(money(raise_total))),
        ("raise_top_classification", top_raise["classification"]),
        ("raise_top_cost", str(top_raise["raise_cost"])),
        ("single_step_classifications",
         str(sum(1 for r in rows if r["steps"] == 1))),
        ("widest_span", str(widest["span"])),
        ("widest_span_classification", widest["classification"]),
    ]


def write_workbook(grid, meta):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    NAVY = "1F3864"
    MIST = "D9E2F3"
    MONEY_FMT = '"$"#,##0.00'
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=NAVY)
    label_font = Font(bold=True, color=NAVY)
    input_fill = PatternFill("solid", fgColor=MIST)

    wb = Workbook()

    # Data sheet: the cleaned long-format grid in progression order.
    ds = wb.active
    ds.title = "Data"
    ds.append(["classification", "step", "step_label", "biweekly_rate"])
    for cell in ds[1]:
        cell.font = header_font
        cell.fill = header_fill
    for plan, steps in grid.items():
        for pos, (label, rate) in enumerate(steps.items(), 1):
            ds.append([plan, pos, label, float(rate)])
    for row in ds.iter_rows(min_row=2, min_col=4, max_col=4):
        row[0].number_format = MONEY_FMT
    ds.column_dimensions["A"].width = 14
    ds.column_dimensions["B"].width = 8
    ds.column_dimensions["C"].width = 12
    ds.column_dimensions["D"].width = 16
    ds.freeze_panes = "A2"

    # Model sheet: pivoted grid plus the live what-if. Step columns are
    # progression positions (Step 1 = lowest rate); portal labels stay
    # on the Data sheet.
    ms = wb.create_sheet("Model")
    max_step = max(len(steps) for steps in grid.values())
    plans = list(grid)
    n_plans = len(plans)

    GRID_HEADER = 18
    first_row = GRID_HEADER + 1
    last_row = GRID_HEADER + n_plans
    step_first_col = 2                       # column B
    step_last_col = 1 + max_step
    d = step_last_col                        # derived columns follow the steps
    col_steps, col_first, col_top = d + 1, d + 2, d + 3
    col_span, col_one, col_base, col_raise = d + 4, d + 5, d + 6, d + 7

    L = get_column_letter
    step_a, step_z = L(step_first_col), L(step_last_col)
    steps_rng = f"{L(col_steps)}{first_row}:{L(col_steps)}{last_row}"
    span_rng = f"{L(col_span)}{first_row}:{L(col_span)}{last_row}"
    base_rng = f"{L(col_base)}{first_row}:{L(col_base)}{last_row}"
    raise_rng = f"{L(col_raise)}{first_row}:{L(col_raise)}{last_row}"
    name_rng = f"A{first_row}:A{last_row}"

    ms["A1"] = "Public pay-band cost model"
    ms["A1"].font = Font(bold=True, size=14, color=NAVY)
    ms["A2"] = (f"NS Government Pay Scales, period starting "
                f"{meta['latest_period_start']}. Published biweekly rates, "
                f"not payroll headcount. Step columns follow each plan's "
                f"pay progression; portal step labels are on the Data "
                f"sheet.")
    ms["A2"].font = Font(italic=True, size=9)

    ms["A4"] = "Inputs"
    ms["A4"].font = label_font
    ms["A5"] = "Across-the-board raise"
    ms["B5"] = float(RAISE_PCT)
    ms["B5"].number_format = "0.0%"
    ms["B5"].fill = input_fill
    ms["B5"].font = Font(bold=True)

    ms["A7"] = "Key figures (per biweekly pay period)"
    ms["A7"].font = label_font
    key_cells = [
        ("Classifications", f"=COUNTA({name_rng})", "0"),
        ("Published step rates", f"=SUM({steps_rng})", "0"),
        ("Total of published biweekly rates", f"=SUM({base_rng})", MONEY_FMT),
        ("Raise cost, all classifications", f"=SUM({raise_rng})", MONEY_FMT),
        ("Costliest classification to raise",
         f"=INDEX({name_rng},MATCH(MAX({raise_rng}),{raise_rng},0))", None),
        ("Its raise cost", f"=MAX({raise_rng})", MONEY_FMT),
        ("Single-step classifications", f"=COUNTIF({steps_rng},1)", "0"),
        ("Widest first-to-top span", f"=MAX({span_rng})", MONEY_FMT),
        ("Classification with the widest span",
         f"=INDEX({name_rng},MATCH(MAX({span_rng}),{span_rng},0))", None),
    ]
    for i, (label, formula, fmt) in enumerate(key_cells):
        r = 8 + i
        ms.cell(row=r, column=1, value=label)
        c = ms.cell(row=r, column=2, value=formula)
        if fmt:
            c.number_format = fmt

    # Grid header.
    ms.cell(row=GRID_HEADER, column=1, value="Classification")
    for s in range(1, max_step + 1):
        ms.cell(row=GRID_HEADER, column=1 + s, value=f"Step {s}")
    for col, label in [
        (col_steps, "Steps"), (col_first, "First step"), (col_top, "Top step"),
        (col_span, "Span"), (col_one, "One-step cost"),
        (col_base, "Base total"), (col_raise, "Raise cost"),
    ]:
        ms.cell(row=GRID_HEADER, column=col, value=label)
    for cell in ms[GRID_HEADER]:
        if cell.value is not None:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    # Grid rows: published rates as values, everything derived as formulas.
    for i, plan in enumerate(plans):
        r = first_row + i
        ms.cell(row=r, column=1, value=plan)
        for pos, rate in enumerate(grid[plan].values(), 1):
            c = ms.cell(row=r, column=1 + pos, value=float(rate))
            c.number_format = MONEY_FMT
        rng = f"{step_a}{r}:{step_z}{r}"
        row_formulas = [
            (col_steps, f"=COUNT({rng})", "0"),
            (col_first, f"=MIN({rng})", MONEY_FMT),
            (col_top, f"=MAX({rng})", MONEY_FMT),
            (col_span, f"={L(col_top)}{r}-{L(col_first)}{r}", MONEY_FMT),
            (col_one,
             f"=IF({L(col_steps)}{r}>1,ROUND({L(col_span)}{r}/"
             f"({L(col_steps)}{r}-1),2),0)", MONEY_FMT),
            (col_base, f"=SUM({rng})", MONEY_FMT),
            (col_raise, f"=SUMPRODUCT(ROUND({rng}*$B$5,2))", MONEY_FMT),
        ]
        for col, formula, fmt in row_formulas:
            c = ms.cell(row=r, column=col, value=formula)
            c.number_format = fmt

    ms.column_dimensions["A"].width = 34
    for col in range(step_first_col, step_last_col + 1):
        ms.column_dimensions[L(col)].width = 12
    for col in range(col_steps, col_raise + 1):
        ms.column_dimensions[L(col)].width = 14
    ms.freeze_panes = f"B{first_row}"

    wb.active = wb.index(ms)
    wb.save(WORKBOOK)


def read_expected():
    if not EXPECTED.exists():
        sys.exit(f"expected file not found: {EXPECTED.relative_to(HERE)}")
    with open(EXPECTED, newline="", encoding="utf-8") as f:
        return [(r["figure"], r["value"]) for r in csv.DictReader(f)]


def write_expected(figures):
    """Maintainer helper: freeze the current figures as the golden file."""
    EXPECTED.parent.mkdir(parents=True, exist_ok=True)
    with open(EXPECTED, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["figure", "value"])
        w.writerows(figures)


def verify(figures):
    expected = read_expected()
    if len(figures) != len(expected):
        print(f"FAIL: {len(figures)} figures computed, "
              f"{len(expected)} expected")
        return False
    for (got_name, got_val), (exp_name, exp_val) in zip(figures, expected):
        if got_name != exp_name or got_val != exp_val:
            print(f"FAIL: first mismatch at figure {exp_name!r}: "
                  f"expected {exp_val!r}, got ({got_name!r}, {got_val!r})")
            return False
    print(f"PASS: all {len(figures)} key figures match "
          f"expected/key_figures.csv")
    return True


def fmt_value(name, value):
    if name in ("total_biweekly_base", "raise_total", "raise_top_cost",
                "widest_span"):
        return f"${Decimal(value):,.2f}"
    if name == "raise_pct":
        return f"{Decimal(value) * 100:.1f}%"
    return value


def show(figures):
    name_w = max(len(n) for n, _ in figures)
    val_w = max(len(fmt_value(n, v)) for n, v in figures)
    print(f"{'figure':<{name_w}}  {'value':>{val_w}}")
    print(f"{'-' * name_w}  {'-' * val_w}")
    for name, value in figures:
        print(f"{name:<{name_w}}  {fmt_value(name, value):>{val_w}}")


def main():
    cmd = sys.argv[1] if len(sys.argv) > 1 else "build"
    grid, meta = load_scale()
    figures = compute_key_figures(grid)

    if cmd == "build":
        write_workbook(grid, meta)
        print(f"wrote {WORKBOOK.name} "
              f"(period {meta['latest_period_start']}, "
              f"{meta['period_rows']} snapshot rows, "
              f"{meta['dropped_blank_rate']} dropped for blank rate, "
              f"{meta['wrapped_plans']} plans with wrapped step numbering)")
        sys.exit(0 if verify(figures) else 1)
    elif cmd == "verify":
        sys.exit(0 if verify(figures) else 1)
    elif cmd == "show":
        show(figures)
    else:
        sys.exit(f"unknown command {cmd!r} (use: build | verify | show)")


if __name__ == "__main__":
    main()
