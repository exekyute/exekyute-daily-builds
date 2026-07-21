"""Build and verify the co-op sector financial health Excel model.

Reads the pinned snapshot in data/raw/, writes coop-sector-financial-health.xlsx
with live formulas on the Model sheet (no VBA, no macros), then recomputes
every key figure in plain Python with decimal.Decimal and ROUND_HALF_UP
(mirroring Excel's ROUND) and diffs the result against
expected/key_figures.csv.

Commands:
    python build.py            rebuild the workbook, then verify
    python build.py verify     key-figure check only
    python build.py show       print the key figures as a table
"""

import csv
import glob
import os
import re
import sys
from decimal import Decimal, ROUND_HALF_UP

HERE = os.path.dirname(os.path.abspath(__file__))
RAW_GLOB = os.path.join(HERE, "data", "raw", "*.csv")
EXPECTED = os.path.join(HERE, "expected", "key_figures.csv")
WORKBOOK = os.path.join(HERE, "coop-sector-financial-health.xlsx")

YEAR_PATTERN = re.compile(r"^\d{4}$")
CENT = Decimal("0.01")

# Snapshot columns the model uses, in Data sheet order. Keys are the
# dataset's own column names; every value must parse as an integer.
COLUMNS = [
    "report_year", "co_ops_reporting", "total_income", "total_expenses",
    "net_income", "total_assets", "total_liability", "total_equity",
    "full_time_employees", "part_time_employees", "total_employees",
    "total_members",
]


def round2(value):
    """Round to two decimals, half away from zero, mirroring Excel ROUND."""
    return value.quantize(CENT, rounding=ROUND_HALF_UP)


def pct(numerator, denominator):
    """numerator / denominator as a percent, rounded to two decimals."""
    return round2(Decimal(numerator) / Decimal(denominator) * 100)


def direction(current, previous):
    """Year-over-year flag on the rounded values: up, down, or flat."""
    if current > previous:
        return "up"
    if current < previous:
        return "down"
    return "flat"


def load_rows():
    """Read the newest snapshot in data/raw and return cleaned rows.

    Cleaning rules (documented in spec.md):
      1. report_year must be a four-digit year.
      2. Every used column must parse as an integer; a failed row is dropped.
      3. Rows with a zero or negative total_income, total_assets,
         total_liability, or co_ops_reporting are dropped, so no formula
         ever divides by zero.
      4. Sort ascending by report_year. A duplicate year stops the build.
    """
    paths = sorted(glob.glob(RAW_GLOB))
    if not paths:
        sys.exit("No snapshot found under data/raw/")
    path = paths[-1]
    rows = []
    with open(path, newline="", encoding="utf-8") as f:
        for rec in csv.DictReader(f):
            year_text = (rec.get("report_year") or "").strip()
            if not YEAR_PATTERN.match(year_text):
                continue
            try:
                row = {name: int(rec[name]) for name in COLUMNS}
            except (KeyError, TypeError, ValueError):
                continue
            if (row["total_income"] <= 0 or row["total_assets"] <= 0
                    or row["total_liability"] <= 0
                    or row["co_ops_reporting"] <= 0):
                continue
            rows.append(row)
    rows.sort(key=lambda r: r["report_year"])
    years = [r["report_year"] for r in rows]
    if len(set(years)) != len(years):
        sys.exit("Snapshot repeats a report_year; refusing to build")
    if not rows:
        sys.exit("Snapshot contained no usable rows")
    return rows


def compute_key_figures(rows):
    """Recompute every key figure the Model sheet holds, in plain Python.

    Returns an ordered list of (figure, value_string) pairs, metric by
    metric, then the sector totals, then the headline block. Direction
    flags compare the rounded ratio values, exactly as the Model sheet's
    IF formulas do.
    """
    years = [r["report_year"] for r in rows]
    margins = [pct(r["net_income"], r["total_income"]) for r in rows]
    equity_ratios = [pct(r["total_equity"], r["total_assets"]) for r in rows]
    solvency = [round2(Decimal(r["total_assets"]) / Decimal(r["total_liability"]))
                for r in rows]
    employees = [round2(Decimal(r["total_employees"])
                        / Decimal(r["co_ops_reporting"])) for r in rows]

    def flags(series):
        return ["n/a"] + [direction(cur, prev)
                          for prev, cur in zip(series, series[1:])]

    margin_dirs = flags(margins)
    equity_dirs = flags(equity_ratios)
    employee_dirs = flags(employees)

    figures = []
    for name, series in [
        ("margin_pct", margins), ("margin_dir", margin_dirs),
        ("equity_ratio_pct", equity_ratios), ("equity_dir", equity_dirs),
        ("solvency_ratio", solvency),
        ("employees_per_coop", employees), ("employees_dir", employee_dirs),
    ]:
        for year, value in zip(years, series):
            figures.append((f"{name}_{year}", f"{value}"))

    total_income = sum(r["total_income"] for r in rows)
    total_expenses = sum(r["total_expenses"] for r in rows)
    total_net = sum(r["net_income"] for r in rows)
    overall_margin = pct(total_net, total_income)
    figures.append(("total_income_all_years", f"{total_income}"))
    figures.append(("total_expenses_all_years", f"{total_expenses}"))
    figures.append(("total_net_income_all_years", f"{total_net}"))
    figures.append(("overall_margin_pct", f"{overall_margin}"))

    figures.append(("latest_year", f"{years[-1]}"))
    figures.append(("latest_margin_pct", f"{margins[-1]}"))
    figures.append(("latest_margin_dir", margin_dirs[-1]))
    figures.append(("latest_equity_ratio_pct", f"{equity_ratios[-1]}"))
    figures.append(("latest_equity_dir", equity_dirs[-1]))
    figures.append(("latest_solvency_ratio", f"{solvency[-1]}"))
    figures.append(("latest_employees_per_coop", f"{employees[-1]}"))
    figures.append(("latest_employees_dir", employee_dirs[-1]))
    return figures


def build_workbook(rows):
    """Write the .xlsx: a Data sheet of prepared values and a Model sheet
    where every key figure is a live formula over the Data sheet."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    bold = Font(bold=True)
    wb = Workbook()

    data = wb.active
    data.title = "Data"
    for col, name in enumerate(COLUMNS, start=1):
        data.cell(row=1, column=col, value=name).font = bold
    for i, row in enumerate(rows, start=2):
        for col, name in enumerate(COLUMNS, start=1):
            cell = data.cell(row=i, column=col, value=row[name])
            if col >= 2:
                cell.number_format = "#,##0"
    for col in range(1, len(COLUMNS) + 1):
        data.column_dimensions[get_column_letter(col)].width = 17
    data.freeze_panes = "A2"

    n = len(rows)
    first = 5                 # first per-year row on the Model sheet
    last = first + n - 1      # last per-year row
    last_data = n + 1         # last row on the Data sheet

    m = wb.create_sheet("Model")
    m["A1"] = "Co-op sector financial health"
    m["A1"].font = Font(bold=True, size=13)
    m["A2"] = (f"Nova Scotia Co-operatives Financial and Operating Summary, "
               f"report years {rows[0]['report_year']} to "
               f"{rows[-1]['report_year']}. Ratios cover the co-ops that "
               f"reported each year.")
    m["A2"].font = Font(italic=True, size=9)

    headers = ["Year", "Total income ($)", "Total expenses ($)",
               "Net income ($)", "Operating margin (%)", "Margin YoY",
               "Equity ratio (%)", "Equity YoY",
               "Solvency (assets / liabilities)",
               "Employees per reporting co-op", "Employees YoY"]
    for col, name in enumerate(headers, start=1):
        m.cell(row=4, column=col, value=name).font = bold

    for i in range(n):
        r = first + i          # Model row
        d = i + 2              # matching Data row
        m.cell(row=r, column=1, value=f"=Data!A{d}")
        m.cell(row=r, column=2, value=f"=Data!C{d}").number_format = "#,##0"
        m.cell(row=r, column=3, value=f"=Data!D{d}").number_format = "#,##0"
        m.cell(row=r, column=4, value=f"=Data!E{d}").number_format = "#,##0"
        m.cell(row=r, column=5,
               value=f"=ROUND(Data!E{d}/Data!C{d}*100,2)"
               ).number_format = "0.00"
        m.cell(row=r, column=7,
               value=f"=ROUND(Data!H{d}/Data!F{d}*100,2)"
               ).number_format = "0.00"
        m.cell(row=r, column=9,
               value=f"=ROUND(Data!F{d}/Data!G{d},2)").number_format = "0.00"
        m.cell(row=r, column=10,
               value=f"=ROUND(Data!K{d}/Data!B{d},2)").number_format = "0.00"
        if i == 0:
            for col in (6, 8, 11):
                m.cell(row=r, column=col, value="n/a")
        else:
            for col, v in ((6, "E"), (8, "G"), (11, "J")):
                m.cell(row=r, column=col,
                       value=(f'=IF({v}{r}>{v}{r - 1},"up",'
                              f'IF({v}{r}<{v}{r - 1},"down","flat"))'))

    t = last + 2
    m.cell(row=t, column=1,
           value=(f"Sector totals, {rows[0]['report_year']} to "
                  f"{rows[-1]['report_year']}")).font = bold
    m.cell(row=t + 1, column=1, value="Total income ($)")
    m.cell(row=t + 1, column=2,
           value=f"=SUM(Data!C2:C{last_data})").number_format = "#,##0"
    m.cell(row=t + 2, column=1, value="Total expenses ($)")
    m.cell(row=t + 2, column=2,
           value=f"=SUM(Data!D2:D{last_data})").number_format = "#,##0"
    m.cell(row=t + 3, column=1, value="Total net income ($)")
    m.cell(row=t + 3, column=2,
           value=f"=SUM(Data!E2:E{last_data})").number_format = "#,##0"
    m.cell(row=t + 4, column=1, value="Overall operating margin (%)")
    m.cell(row=t + 4, column=2,
           value=(f"=ROUND(SUM(Data!E2:E{last_data})"
                  f"/SUM(Data!C2:C{last_data})*100,2)")
           ).number_format = "0.00"

    h = t + 6
    m.cell(row=h, column=1, value="Headline").font = bold
    headline = [
        ("Latest report year", f"=A{last}", None),
        ("Operating margin (%)", f"=E{last}", "0.00"),
        ("Margin direction", f"=F{last}", None),
        ("Equity ratio (%)", f"=G{last}", "0.00"),
        ("Equity ratio direction", f"=H{last}", None),
        ("Solvency (assets / liabilities)", f"=I{last}", "0.00"),
        ("Employees per reporting co-op", f"=J{last}", "0.00"),
        ("Employees direction", f"=K{last}", None),
    ]
    for i, (label, formula, fmt) in enumerate(headline, start=1):
        m.cell(row=h + i, column=1, value=label)
        cell = m.cell(row=h + i, column=2, value=formula)
        if fmt:
            cell.number_format = fmt

    m.column_dimensions["A"].width = 32
    for col in range(2, 12):
        m.column_dimensions[get_column_letter(col)].width = 17
    m.freeze_panes = f"A{first}"
    wb.active = m
    wb.save(WORKBOOK)


def write_expected(figures):
    """Maintainer helper: freeze the current figures as the golden file."""
    os.makedirs(os.path.dirname(EXPECTED), exist_ok=True)
    with open(EXPECTED, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["figure", "value"])
        w.writerows(figures)


def verify():
    """Recompute every key figure and diff against expected/key_figures.csv."""
    computed = compute_key_figures(load_rows())
    if not os.path.exists(EXPECTED):
        sys.exit("expected/key_figures.csv is missing")
    expected = []
    with open(EXPECTED, newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        header = next(reader, None)
        if header != ["figure", "value"]:
            sys.exit("expected/key_figures.csv has an unexpected header")
        for rec in reader:
            if rec:
                expected.append((rec[0], rec[1]))
    if len(computed) != len(expected):
        print(f"FAIL: {len(computed)} computed figures, "
              f"{len(expected)} expected")
        sys.exit(1)
    for (cf, cv), (ef, ev) in zip(computed, expected):
        if cf != ef or cv != ev:
            print(f"FAIL: first mismatch at {ef}: expected {ev}, got "
                  f"{cv} (computed name {cf})")
            sys.exit(1)
    print(f"PASS: all {len(computed)} key figures match "
          "expected/key_figures.csv")


def show():
    """Print the key figures as an aligned plain-ASCII table."""
    computed = compute_key_figures(load_rows())
    name_width = max(len(name) for name, _ in computed)
    value_width = max(len(value) for _, value in computed)
    print(f"{'figure'.ljust(name_width)}  {'value'.rjust(value_width)}")
    print(f"{'-' * name_width}  {'-' * value_width}")
    for name, value in computed:
        print(f"{name.ljust(name_width)}  {value.rjust(value_width)}")


def main():
    command = sys.argv[1] if len(sys.argv) > 1 else "build"
    if command == "build":
        rows = load_rows()
        build_workbook(rows)
        print(f"Wrote {os.path.basename(WORKBOOK)} "
              f"({len(rows)} report years)")
        verify()
    elif command == "verify":
        verify()
    elif command == "show":
        show()
    else:
        sys.exit("Usage: python build.py [verify|show]")


if __name__ == "__main__":
    main()
