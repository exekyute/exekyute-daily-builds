"""Build and verify the children's dental utilization Excel model.

Reads the pinned snapshot in data/raw/, writes childrens-dental-utilization.xlsx
with live formulas on the Model sheet, then recomputes every key figure in
plain Python and diffs the result against expected/key_figures.csv.

Commands:
    python build.py            rebuild the workbook, then verify
    python build.py verify     key-figure check only
    python build.py show       print the key figures as a table
"""

import csv
import glob
import os
import re
import sys
from decimal import Decimal, ROUND_HALF_UP

HERE = os.path.dirname(os.path.abspath(__file__))
RAW_GLOB = os.path.join(HERE, "data", "raw", "*.csv")
EXPECTED = os.path.join(HERE, "expected", "key_figures.csv")
WORKBOOK = os.path.join(HERE, "childrens-dental-utilization.xlsx")

FY_PATTERN = re.compile(r"^\d{4}-\d{4}$")
CENT = Decimal("0.01")
SIXDP = Decimal("0.000001")


def money(value):
    """Round to the cent, half away from zero, mirroring Excel ROUND."""
    return value.quantize(CENT, rounding=ROUND_HALF_UP)


def load_rows():
    """Read the newest snapshot in data/raw and return cleaned rows.

    Each row: (fiscal_year, services, amount_paid, insured, beneficiaries).
    Rows with a malformed fiscal year, missing counts, or zero
    beneficiaries / insured persons are dropped. Sorted by fiscal year.
    """
    paths = sorted(glob.glob(RAW_GLOB))
    if not paths:
        sys.exit("No snapshot found under data/raw/")
    path = paths[-1]
    rows = []
    with open(path, newline="", encoding="utf-8") as f:
        for rec in csv.DictReader(f):
            fy = (rec.get("fiscal_year") or "").strip()
            if not FY_PATTERN.match(fy):
                continue
            try:
                services = int(rec["_of_services_rendered"])
                paid = Decimal(rec["amount_paid"])
                insured = int(rec["persons_insured"])
                benef = int(rec["beneficiaries"])
            except (KeyError, ValueError, ArithmeticError):
                continue
            if benef <= 0 or insured <= 0:
                continue
            rows.append((fy, services, paid, insured, benef))
    rows.sort(key=lambda r: r[0])
    if not rows:
        sys.exit("Snapshot contained no usable rows")
    return rows


def compute_key_figures(rows):
    """Recompute every key figure the Model sheet holds, in plain Python.

    Returns an ordered list of (figure, value_string) pairs. The projection
    uses the closed-form least squares slope and intercept (covariance over
    variance), never a fitting library.
    """
    figures = []
    years = [int(fy[:4]) for fy, *_ in rows]
    ppb = []  # paid per beneficiary, rounded to the cent
    for fy, _services, paid, insured, benef in rows:
        value = money(paid / Decimal(benef))
        ppb.append(value)
        figures.append((f"paid_per_beneficiary_{fy}", f"{value}"))
    for fy, _services, _paid, insured, benef in rows:
        rate = (Decimal(benef) / Decimal(insured) * 100).quantize(
            CENT, rounding=ROUND_HALF_UP
        )
        figures.append((f"coverage_rate_pct_{fy}", f"{rate}"))

    total_paid = sum(paid for _fy, _s, paid, _i, _b in rows)
    total_benef = sum(benef for _fy, _s, _p, _i, benef in rows)
    overall = money(total_paid / Decimal(total_benef))
    figures.append(("total_amount_paid", f"{total_paid}"))
    figures.append(("total_beneficiaries", f"{total_benef}"))
    figures.append(("overall_paid_per_beneficiary", f"{overall}"))

    # Least squares over the observed years: slope = covariance / variance.
    xs = [float(y) for y in years]
    ys = [float(v) for v in ppb]
    n = len(xs)
    mean_x = sum(xs) / n
    mean_y = sum(ys) / n
    cov = sum((x - mean_x) * (y - mean_y) for x, y in zip(xs, ys))
    var = sum((x - mean_x) ** 2 for x in xs)
    slope = cov / var
    intercept = mean_y - slope * mean_x
    figures.append(
        ("slope_per_year",
         f"{Decimal(repr(slope)).quantize(SIXDP, rounding=ROUND_HALF_UP)}")
    )
    figures.append(
        ("intercept",
         f"{Decimal(repr(intercept)).quantize(SIXDP, rounding=ROUND_HALF_UP)}")
    )

    last_year = years[-1]
    for step in (1, 2):
        target = last_year + step
        label = f"{target}-{target + 1}"
        projected = Decimal(repr(intercept + slope * target)).quantize(
            CENT, rounding=ROUND_HALF_UP
        )
        figures.append((f"projected_paid_per_beneficiary_{label}",
                        f"{projected}"))

    latest = ppb[-1]
    first = ppb[0]
    change = ((latest - first) / first * 100).quantize(
        CENT, rounding=ROUND_HALF_UP
    )
    figures.append(("latest_paid_per_beneficiary", f"{latest}"))
    figures.append(("change_pct_first_to_latest", f"{change}"))
    return figures


def build_workbook(rows):
    """Write the .xlsx: a Data sheet of prepared values and a Model sheet
    where every key figure is a live formula over the Data sheet."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    bold = Font(bold=True)
    wb = Workbook()

    data = wb.active
    data.title = "Data"
    headers = ["fiscal_year", "services_rendered", "amount_paid",
               "persons_insured", "beneficiaries"]
    for col, name in enumerate(headers, start=1):
        cell = data.cell(row=1, column=col, value=name)
        cell.font = bold
    for i, (fy, services, paid, insured, benef) in enumerate(rows, start=2):
        data.cell(row=i, column=1, value=fy)
        data.cell(row=i, column=2, value=services)
        paid_cell = data.cell(
            row=i, column=3,
            value=int(paid) if paid == paid.to_integral_value() else float(paid)
        )
        paid_cell.number_format = "#,##0"
        data.cell(row=i, column=4, value=insured).number_format = "#,##0"
        data.cell(row=i, column=5, value=benef).number_format = "#,##0"
    for col in range(1, 6):
        data.column_dimensions[get_column_letter(col)].width = 18

    n = len(rows)
    last_data = n + 1          # last row on the Data sheet
    start = 4                  # first per-year row on the Model sheet
    end = start + n - 1        # last per-year row on the Model sheet

    m = wb.create_sheet("Model")
    m["A1"] = "Children's dental utilization model"
    m["A1"].font = Font(bold=True, size=13)

    table_headers = ["Fiscal year", "Year start", "Amount paid ($)",
                     "Beneficiaries", "Paid per beneficiary ($)",
                     "Persons insured", "Coverage rate (%)"]
    for col, name in enumerate(table_headers, start=1):
        cell = m.cell(row=3, column=col, value=name)
        cell.font = bold
    for i in range(n):
        r = start + i          # Model row
        d = i + 2              # matching Data row
        m.cell(row=r, column=1, value=f"=Data!A{d}")
        m.cell(row=r, column=2, value=f"=VALUE(LEFT(Data!A{d},4))")
        m.cell(row=r, column=3,
               value=f"=Data!C{d}").number_format = "#,##0"
        m.cell(row=r, column=4,
               value=f"=Data!E{d}").number_format = "#,##0"
        m.cell(row=r, column=5,
               value=f"=ROUND(Data!C{d}/Data!E{d},2)").number_format = "0.00"
        m.cell(row=r, column=6,
               value=f"=Data!D{d}").number_format = "#,##0"
        m.cell(row=r, column=7,
               value=f"=ROUND(Data!E{d}/Data!D{d}*100,2)").number_format = "0.00"

    t = end + 2
    m.cell(row=t, column=1, value="Totals").font = bold
    m.cell(row=t + 1, column=1, value="Total amount paid ($)")
    m.cell(row=t + 1, column=2,
           value=f"=SUM(Data!C2:C{last_data})").number_format = "#,##0"
    m.cell(row=t + 2, column=1, value="Total beneficiaries")
    m.cell(row=t + 2, column=2,
           value=f"=SUM(Data!E2:E{last_data})").number_format = "#,##0"
    m.cell(row=t + 3, column=1, value="Overall paid per beneficiary ($)")
    m.cell(
        row=t + 3, column=2,
        value=f"=ROUND(SUM(Data!C2:C{last_data})/SUM(Data!E2:E{last_data}),2)"
    ).number_format = "0.00"

    y_range = f"$E${start}:$E${end}"
    x_range = f"$B${start}:$B${end}"
    p = t + 5
    m.cell(row=p, column=1,
           value="Trend and projection (least squares over observed years)"
           ).font = bold
    m.cell(row=p + 1, column=1, value="Slope ($ per year)")
    m.cell(row=p + 1, column=2,
           value=f"=ROUND(SLOPE({y_range},{x_range}),6)"
           ).number_format = "0.000000"
    m.cell(row=p + 2, column=1, value="Intercept ($)")
    m.cell(row=p + 2, column=2,
           value=f"=ROUND(INTERCEPT({y_range},{x_range}),6)"
           ).number_format = "0.000000"
    for col, name in enumerate(
            ["Projected period", "Year start",
             "Projected paid per beneficiary ($)"], start=1):
        m.cell(row=p + 3, column=col, value=name).font = bold
    for step in (1, 2):
        r = p + 3 + step
        m.cell(row=r, column=1,
               value=f'=TEXT(B{r},"0")&"-"&TEXT(B{r}+1,"0")')
        m.cell(row=r, column=2, value=f"=B{end}+{step}")
        m.cell(
            row=r, column=3,
            value=(f"=ROUND(INTERCEPT({y_range},{x_range})"
                   f"+SLOPE({y_range},{x_range})*B{r},2)")
        ).number_format = "0.00"

    h = p + 7
    m.cell(row=h, column=1, value="Headline").font = bold
    m.cell(row=h + 1, column=1, value="Latest fiscal year")
    m.cell(row=h + 1, column=2, value=f"=A{end}")
    m.cell(row=h + 2, column=1, value="Latest paid per beneficiary ($)")
    m.cell(row=h + 2, column=2, value=f"=E{end}").number_format = "0.00"
    m.cell(row=h + 3, column=1, value="Change since first year (%)")
    m.cell(row=h + 3, column=2,
           value=f"=ROUND((E{end}-E{start})/E{start}*100,2)"
           ).number_format = "0.00"
    m.cell(row=h + 4, column=1,
           value="Next projected paid per beneficiary ($)")
    m.cell(row=h + 4, column=2, value=f"=C{p + 4}").number_format = "0.00"

    m.column_dimensions["A"].width = 36
    for col in range(2, 8):
        m.column_dimensions[get_column_letter(col)].width = 16

    wb.save(WORKBOOK)


def verify():
    """Recompute every key figure and diff against expected/key_figures.csv."""
    computed = compute_key_figures(load_rows())
    if not os.path.exists(EXPECTED):
        sys.exit("expected/key_figures.csv is missing")
    expected = []
    with open(EXPECTED, newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        header = next(reader, None)
        if header != ["figure", "value"]:
            sys.exit("expected/key_figures.csv has an unexpected header")
        for rec in reader:
            if rec:
                expected.append((rec[0], rec[1]))
    if len(computed) != len(expected):
        print(f"FAIL: {len(computed)} computed figures, "
              f"{len(expected)} expected")
        sys.exit(1)
    for (cf, cv), (ef, ev) in zip(computed, expected):
        if cf != ef or cv != ev:
            print(f"FAIL: first mismatch at {ef}: expected {ev}, got "
                  f"{cv} (computed name {cf})")
            sys.exit(1)
    print(f"PASS: all {len(computed)} key figures match "
          "expected/key_figures.csv")


def show():
    """Print the key figures as an aligned plain-ASCII table."""
    computed = compute_key_figures(load_rows())
    name_width = max(len(name) for name, _ in computed)
    value_width = max(len(value) for _, value in computed)
    print(f"{'figure'.ljust(name_width)}  {'value'.rjust(value_width)}")
    print(f"{'-' * name_width}  {'-' * value_width}")
    for name, value in computed:
        print(f"{name.ljust(name_width)}  {value.rjust(value_width)}")


def main():
    command = sys.argv[1] if len(sys.argv) > 1 else "build"
    if command == "build":
        rows = load_rows()
        build_workbook(rows)
        print(f"Wrote {os.path.basename(WORKBOOK)} "
              f"({len(rows)} fiscal years)")
        verify()
    elif command == "verify":
        verify()
    elif command == "show":
        show()
    else:
        sys.exit("Usage: python build.py [verify|show]")


if __name__ == "__main__":
    main()
