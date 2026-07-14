"""Highway capital-plan tracker: builds and verifies the Excel model.

Commands:
    python build.py          rebuild the workbook, then verify key figures
    python build.py verify   re-run the key-figure check only
    python build.py show     print the key figures as an aligned table

The workbook is formula-driven: every key figure on the Model sheet is a live
SUMIFS/COUNTIFS (or INDEX/MATCH) formula over the Data sheet. This script
recomputes the same figures in plain Python from the raw snapshot and diffs
them against expected/key_figures.csv.

Rounding note: kilometre figures are rounded half-away-from-zero with
decimal.Decimal and ROUND_HALF_UP, matching Excel's ROUND function. Python's
built-in round() rounds half to even and is never used for these figures.
"""

import csv
import sys
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

HERE = Path(__file__).resolve().parent
RAW_DIR = HERE / "data" / "raw"
EXPECTED_PATH = HERE / "expected" / "key_figures.csv"
XLSX_PATH = HERE / "highway-capital-plan-tracker.xlsx"

CENT = Decimal("0.01")
SHARE_Q = Decimal("0.0001")


# ---------------------------------------------------------------- data layer

def find_snapshot():
    files = sorted(RAW_DIR.glob("ns_highway-improvement-plan_*.csv"))
    if not files:
        sys.exit(f"no snapshot found in {RAW_DIR.relative_to(HERE)}")
    return files[-1]


def load_rows(path):
    """Read the snapshot and apply the cleaning rules from spec.md."""
    rows = []
    with open(path, newline="", encoding="utf-8") as f:
        for rec in csv.DictReader(f):
            source = rec["source"].strip()
            county = rec["county"].strip() or "Unspecified"
            ptype = rec["construct_"].strip() or "Unspecified"
            if ptype == "Gravel Roads Program":
                ptype = "Gravel Road Program"
            year = rec["year_start"].strip() or "Unknown"
            status = rec["status"].strip() or "Unspecified"
            km_raw = rec.get("km", "").strip()
            km = Decimal(km_raw) if km_raw else None
            rows.append({
                "source": source,
                "project": rec["project_de"].strip(),
                "county": county,
                "type": ptype,
                "year": year,
                "km": km,
                "status": status,
            })
    return rows


# ------------------------------------------------------------- computations

def sorted_unique(rows, field):
    return sorted({r[field] for r in rows})


def compute_figures(rows):
    """Recompute every Model-sheet key figure in plain Python.

    Returns an ordered list of (metric, key, value) string triples, the same
    shape as expected/key_figures.csv.
    """
    counties = sorted_unique(rows, "county")
    years = sorted_unique(rows, "year")
    types = sorted_unique(rows, "type")

    total = len(rows)
    by_source = {s: sum(1 for r in rows if r["source"] == s)
                 for s in sorted_unique(rows, "source")}
    by_county = {c: sum(1 for r in rows if r["county"] == c) for c in counties}
    by_year = {y: sum(1 for r in rows if r["year"] == y) for y in years}
    by_type = {t: sum(1 for r in rows if r["type"] == t) for t in types}

    # Leading county / top type: first occurrence of the maximum in
    # alphabetical order, matching INDEX/MATCH(MAX(...)) on the Model sheet.
    lead_county = next(c for c in counties if by_county[c] == max(by_county.values()))
    top_type = next(t for t in types if by_type[t] == max(by_type.values()))

    def share(n):
        return (Decimal(n) / Decimal(total)).quantize(SHARE_Q, ROUND_HALF_UP)

    # Road kilometres by fiscal year: each per-year figure is rounded to two
    # decimals, and the total is the sum of the rounded figures so the block
    # ties exactly, mirroring the SUM over ROUND(...) cells in the workbook.
    km_by_year = {}
    for y in years:
        vals = [r["km"] for r in rows
                if r["year"] == y and r["source"] == "roads" and r["km"] is not None]
        km_by_year[y] = sum(vals, Decimal(0)).quantize(CENT, ROUND_HALF_UP)
    km_total = sum(km_by_year.values(), Decimal(0)).quantize(CENT, ROUND_HALF_UP)

    fig = [("total_projects", "", str(total))]
    for s, n in by_source.items():
        fig.append(("projects_by_source", s, str(n)))
    fig.append(("leading_county", "", lead_county))
    fig.append(("leading_county_projects", "", str(by_county[lead_county])))
    fig.append(("top_type", "", top_type))
    fig.append(("top_type_share", "", str(share(by_type[top_type]))))
    for c in counties:
        fig.append(("projects_by_county", c, str(by_county[c])))
    for y in years:
        fig.append(("projects_by_year", y, str(by_year[y])))
    for t in types:
        fig.append(("type_share_overall", t, str(share(by_type[t]))))
    for y in years:
        fig.append(("road_km_by_year", y, str(km_by_year[y])))
    fig.append(("road_km_total", "", str(km_total)))
    return fig


# ------------------------------------------------------------ workbook layer

def write_workbook(rows):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    counties = sorted_unique(rows, "county")
    years = sorted_unique(rows, "year")
    types = sorted_unique(rows, "type")
    n = len(rows)
    last = n + 1  # last data row (header on row 1)

    wb = Workbook()
    wb.calculation.fullCalcOnLoad = True

    # ---- Data sheet: cleaned, typed rows
    ds = wb.active
    ds.title = "Data"
    headers = ["source", "project", "county", "type", "year", "km", "status"]
    ds.append(headers)
    for r in rows:
        ds.append([r["source"], r["project"], r["county"], r["type"],
                   r["year"], float(r["km"]) if r["km"] is not None else None,
                   r["status"]])

    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="1F3B57")
    for c in range(1, len(headers) + 1):
        cell = ds.cell(row=1, column=c)
        cell.font = head_font
        cell.fill = head_fill
    widths = [9, 58, 16, 44, 12, 9, 22]
    for i, w in enumerate(widths, start=1):
        ds.column_dimensions[get_column_letter(i)].width = w
    ds.freeze_panes = "A2"

    # Fixed ranges the Model formulas point at
    rng = {
        "source": f"Data!$A$2:$A${last}",
        "county": f"Data!$C$2:$C${last}",
        "type": f"Data!$D$2:$D${last}",
        "year": f"Data!$E$2:$E${last}",
        "km": f"Data!$F$2:$F${last}",
    }

    ms = wb.create_sheet("Model")
    title_font = Font(bold=True, size=13, color="1F3B57")
    block_font = Font(bold=True, color="1F3B57")
    label_font = Font(bold=True)
    band_fill = PatternFill("solid", fgColor="DCE6F0")
    pct = "0.00%"
    two = "0.00"

    def block_header(row, text):
        cell = ms.cell(row=row, column=1, value=text)
        cell.font = block_font
        cell.fill = band_fill

    ms["A1"] = "Highway capital-plan tracker"
    ms["A1"].font = title_font

    # ---- Headline block
    block_header(3, "Headline")
    cellmap = {}

    def headline(row, label, formula, fmt=None):
        ms.cell(row=row, column=1, value=label).font = label_font
        cell = ms.cell(row=row, column=2, value=formula)
        if fmt:
            cell.number_format = fmt
        return f"B{row}"

    # County totals live in the county-by-year matrix below; formulas here
    # reference that block so the headline stays live.
    mat_top = 12                     # first county row of the matrix
    mat_head = mat_top - 1           # matrix header row
    mat_bot = mat_top + len(counties) - 1
    tot_col = 2 + len(years)         # matrix row-total column index
    tot_l = get_column_letter(tot_col)
    county_names = f"$A${mat_top}:$A${mat_bot}"
    county_tots = f"${tot_l}${mat_top}:${tot_l}${mat_bot}"

    # Type mix block position (after the matrix)
    mix_top = mat_bot + 4
    mix_head = mix_top - 1
    mix_bot = mix_top + len(types) - 1
    mix_tot_col = 2 + len(years)
    mix_tot_l = get_column_letter(mix_tot_col)
    mix_share_col = mix_tot_col + 1
    mix_share_l = get_column_letter(mix_share_col)
    type_names = f"$A${mix_top}:$A${mix_bot}"
    type_tots = f"${mix_tot_l}${mix_top}:${mix_tot_l}${mix_bot}"
    type_shares = f"${mix_share_l}${mix_top}:${mix_share_l}${mix_bot}"

    # Road km block position (after the type mix)
    km_top = mix_bot + 4
    km_head = km_top - 1

    cellmap["total_projects"] = headline(
        4, "Total planned projects", f"=COUNTA({rng['source']})")
    cellmap["leading_county"] = headline(
        5, "Leading county",
        f"=INDEX({county_names},MATCH(MAX({county_tots}),{county_tots},0))")
    cellmap["leading_county_projects"] = headline(
        6, "Projects in leading county", f"=MAX({county_tots})")
    cellmap["top_type"] = headline(
        7, "Largest project type",
        f"=INDEX({type_names},MATCH(MAX({type_tots}),{type_tots},0))")
    cellmap["top_type_share"] = headline(
        8, "Largest type share of all projects",
        f"=MAX({type_shares})", pct)
    cellmap["projects_by_source"] = headline(
        9, "Road / bridge split",
        f'=COUNTIFS({rng["source"]},"roads")&" roads / "'
        f'&COUNTIFS({rng["source"]},"bridges")&" bridges"')

    # ---- County x fiscal-year matrix (project counts)
    block_header(mat_head - 1, "Projects by county and fiscal year")
    ms.cell(row=mat_head, column=1, value="county").font = label_font
    for j, y in enumerate(years):
        ms.cell(row=mat_head, column=2 + j, value=y).font = label_font
    ms.cell(row=mat_head, column=tot_col, value="total").font = label_font
    for i, county in enumerate(counties):
        r = mat_top + i
        ms.cell(row=r, column=1, value=county)
        for j, y in enumerate(years):
            col_l = get_column_letter(2 + j)
            ms.cell(row=r, column=2 + j,
                    value=f"=COUNTIFS({rng['county']},$A{r},"
                          f"{rng['year']},{col_l}${mat_head})")
        ms.cell(row=r, column=tot_col,
                value=f"=SUM(B{r}:{get_column_letter(tot_col - 1)}{r})")
    tr = mat_bot + 1
    ms.cell(row=tr, column=1, value="total").font = label_font
    for j in range(len(years) + 1):
        col_l = get_column_letter(2 + j)
        ms.cell(row=tr, column=2 + j,
                value=f"=SUM({col_l}{mat_top}:{col_l}{mat_bot})").font = label_font

    # ---- Project-type mix by fiscal year (counts plus share of all projects)
    block_header(mix_head - 1, "Project-type mix by fiscal year")
    ms.cell(row=mix_head, column=1, value="project type").font = label_font
    for j, y in enumerate(years):
        ms.cell(row=mix_head, column=2 + j, value=y).font = label_font
    ms.cell(row=mix_head, column=mix_tot_col, value="total").font = label_font
    ms.cell(row=mix_head, column=mix_share_col, value="share").font = label_font
    for i, t in enumerate(types):
        r = mix_top + i
        ms.cell(row=r, column=1, value=t)
        for j, y in enumerate(years):
            col_l = get_column_letter(2 + j)
            ms.cell(row=r, column=2 + j,
                    value=f"=COUNTIFS({rng['type']},$A{r},"
                          f"{rng['year']},{col_l}${mix_head})")
        ms.cell(row=r, column=mix_tot_col,
                value=f"=SUM(B{r}:{get_column_letter(mix_tot_col - 1)}{r})")
        sc = ms.cell(row=r, column=mix_share_col,
                     value=f"=ROUND({mix_tot_l}{r}/COUNTA({rng['source']}),4)")
        sc.number_format = pct
    mr = mix_bot + 1
    ms.cell(row=mr, column=1, value="total").font = label_font
    for j in range(len(years) + 1):
        col_l = get_column_letter(2 + j)
        ms.cell(row=mr, column=2 + j,
                value=f"=SUM({col_l}{mix_top}:{col_l}{mix_bot})").font = label_font

    # ---- Road kilometres by fiscal year
    block_header(km_head - 1, "Road kilometres by fiscal year (roads only)")
    ms.cell(row=km_head, column=1, value="fiscal year").font = label_font
    ms.cell(row=km_head, column=2, value="km").font = label_font
    for i, y in enumerate(years):
        r = km_top + i
        ms.cell(row=r, column=1, value=y)
        kc = ms.cell(row=r, column=2,
                     value=f"=ROUND(SUMIFS({rng['km']},{rng['year']},$A{r},"
                           f'{rng["source"]},"roads"),2)')
        kc.number_format = two
    kr = km_top + len(years)
    ms.cell(row=kr, column=1, value="total").font = label_font
    kt = ms.cell(row=kr, column=2, value=f"=SUM(B{km_top}:B{kr - 1})")
    kt.number_format = two
    kt.font = label_font

    ms.column_dimensions["A"].width = 46
    for c in range(2, mix_share_col + 1):
        ms.column_dimensions[get_column_letter(c)].width = 12

    wb.save(XLSX_PATH)

    cellmap.update({
        "county_matrix": f"B{mat_top}:{get_column_letter(tot_col)}{mat_bot}",
        "type_mix": f"B{mix_top}:{mix_share_l}{mix_bot}",
        "km_block": f"B{km_top}:B{kr}",
    })
    return cellmap


# ---------------------------------------------------------- verify and show

def read_expected():
    if not EXPECTED_PATH.exists():
        return None
    with open(EXPECTED_PATH, newline="", encoding="utf-8") as f:
        return [(r["metric"], r["key"], r["value"]) for r in csv.DictReader(f)]


def verify(rows):
    actual = compute_figures(rows)
    expected = read_expected()
    if expected is None:
        sys.exit(f"missing {EXPECTED_PATH.relative_to(HERE)}; "
                 "generate it from a verified run first")
    if len(actual) != len(expected):
        print(f"FAIL: {len(actual)} figures computed, "
              f"{len(expected)} expected")
        return False
    for a, e in zip(actual, expected):
        if a != e:
            print(f"FAIL: first mismatch at metric={e[0]} key={e[1]}: "
                  f"expected {e[2]}, got {a[2]}")
            return False
    print(f"PASS: all {len(actual)} key figures match "
          f"{EXPECTED_PATH.relative_to(HERE)}")
    return True


def show(rows):
    figures = compute_figures(rows)
    w_m = max(len(f[0]) for f in figures)
    w_k = max(len(f[1]) for f in figures)
    w_v = max(len(f[2]) for f in figures)
    print(f"{'metric'.ljust(w_m)}  {'key'.ljust(w_k)}  {'value'.rjust(w_v)}")
    print(f"{'-' * w_m}  {'-' * w_k}  {'-' * w_v}")
    for m, k, v in figures:
        print(f"{m.ljust(w_m)}  {k.ljust(w_k)}  {v.rjust(w_v)}")


def main():
    cmd = sys.argv[1] if len(sys.argv) > 1 else "build"
    snapshot = find_snapshot()
    rows = load_rows(snapshot)
    if cmd == "build":
        write_workbook(rows)
        print(f"wrote {XLSX_PATH.name} "
              f"({len(rows)} data rows from {snapshot.name})")
        sys.exit(0 if verify(rows) else 1)
    elif cmd == "verify":
        sys.exit(0 if verify(rows) else 1)
    elif cmd == "show":
        show(rows)
    else:
        sys.exit(f"unknown command: {cmd} (use build, verify, or show)")


if __name__ == "__main__":
    main()
