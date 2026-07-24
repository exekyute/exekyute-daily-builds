"""Build and verify the climate-projection summary Excel model.

Reads the pinned snapshot in data/raw/, writes climate-projection-summary.xlsx
with live formulas on the Model sheet, then recomputes every key figure in
plain Python and diffs the result against expected/key_figures.csv.

Commands:
    python build.py            rebuild the workbook, then verify
    python build.py verify     key-figure check only
    python build.py show       print the key figures as a table

Rounding note: figures are rounded half-away-from-zero with decimal.Decimal
and ROUND_HALF_UP, matching Excel's ROUND function. Python's built-in round()
rounds half to even and is never used for reported figures.
"""

import csv
import glob
import os
import sys
from decimal import Decimal, ROUND_HALF_UP

HERE = os.path.dirname(os.path.abspath(__file__))
RAW_GLOB = os.path.join(HERE, "data", "raw", "*.csv")
EXPECTED = os.path.join(HERE, "expected", "key_figures.csv")
WORKBOOK = os.path.join(HERE, "climate-projection-summary.xlsx")

TWODP = Decimal("0.01")

# The model uses one variable at the model-range median (p50): annual mean
# temperature. The 2010 column is the dataset's 1981-2010 baseline period.
VARIABLE = "tgmean_annual"
PROVINCE_ROW = "Nova Scotia"
BASELINE = "2010"
HORIZONS = ("2045", "2095")
PERIODS = (BASELINE,) + HORIZONS
SCENARIOS = ("RCP4.5", "RCP8.5")
SOURCE_COLS = {"RCP4.5": "rcp45", "RCP8.5": "rcp85"}


def two_dp(value):
    """Round to two decimals, half away from zero, mirroring Excel ROUND."""
    return Decimal(repr(float(value))).quantize(TWODP, rounding=ROUND_HALF_UP)


def load_raw():
    """Read the newest snapshot in data/raw as a list of dicts."""
    paths = sorted(glob.glob(RAW_GLOB))
    if not paths:
        sys.exit("No snapshot found under data/raw/")
    with open(paths[-1], newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def prepare_rows(raw):
    """Normalize the snapshot into canonical long rows.

    Keeps the annual-mean-temperature rows for the 18 counties (the
    province-wide 'Nova Scotia' row is dropped so the ranking compares
    counties with counties) and unpivots the p50 columns into one row per
    (region, scenario, period): ('Halifax', 'RCP4.5', '2045', 6.79...).
    """
    rows = []
    for rec in raw:
        if (rec.get("variable") or "").strip() != VARIABLE:
            continue
        region = (rec.get("region") or "").strip()
        if not region or region == PROVINCE_ROW:
            continue
        for scenario in SCENARIOS:
            prefix = SOURCE_COLS[scenario]
            for period in PERIODS:
                cell = (rec.get(f"{prefix}_p50_{period}") or "").strip()
                try:
                    value = float(cell)
                except ValueError:
                    continue
                rows.append((region, scenario, period, value))
    rows.sort(key=lambda r: (r[0], r[1], r[2]))
    if not rows:
        sys.exit(f"Snapshot contained no usable {VARIABLE} rows")
    return rows


def pivot(rows):
    """Aggregate canonical rows into per-region figures.

    Returns (regions, table) where regions is the sorted region list and
    table maps region -> dict with keys '<scenario>_<period>' (mean value
    rounded to two decimals as Decimal, or None when the region has no rows
    for that slot) and 'delta_<scenario>_<horizon>' (projection minus the
    same scenario's baseline).
    """
    buckets = {}
    for region, scenario, period, value in rows:
        key = f"{scenario}_{period}"
        buckets.setdefault(region, {}).setdefault(key, []).append(value)
    regions = sorted(buckets)
    table = {}
    for region in regions:
        cells = {}
        for scenario in SCENARIOS:
            for period in PERIODS:
                key = f"{scenario}_{period}"
                vals = buckets[region].get(key)
                cells[key] = two_dp(sum(vals) / len(vals)) if vals else None
        for scenario in SCENARIOS:
            base = cells[f"{scenario}_{BASELINE}"]
            for horizon in HORIZONS:
                proj = cells[f"{scenario}_{horizon}"]
                cells[f"delta_{scenario}_{horizon}"] = (
                    two_dp(proj - base)
                    if proj is not None and base is not None else None
                )
        table[region] = cells
    return regions, table


def slug(region):
    """Region label as a key-figure suffix: lower case, spaces to hyphens."""
    return region.strip().lower().replace(" ", "-")


def short(scenario):
    """Scenario label as a key-figure fragment: RCP4.5 -> rcp45."""
    return scenario.replace("RCP", "rcp").replace(".", "")


def sort_key(delta, position, n):
    """Tie-safe integer sort key for the ranked block.

    Mirrors the Model sheet's sort-key column (L): the two-decimal delta as an
    integer number of hundredths, shifted three digits, plus the region's
    reverse pivot position, so equal deltas rank the alphabetically earlier
    region first and every region gets exactly one rank.
    """
    return int(delta * 100) * 1000 + (n - 1 - position)


def ranked_regions(regions, table):
    """Regions ranked by 2095 RCP8.5 delta, largest first.

    Mirrors the Model sheet's LARGE + INDEX/MATCH block over the sort-key
    column: rank order is delta descending, ties broken by pivot
    (alphabetical) order.
    """
    n = len(regions)
    keyed = [(sort_key(table[r]["delta_RCP8.5_2095"], i, n), r,
              table[r]["delta_RCP8.5_2095"])
             for i, r in enumerate(regions)
             if table[r]["delta_RCP8.5_2095"] is not None]
    keyed.sort(reverse=True)
    return [(k, region, delta)
            for k, (_key, region, delta) in enumerate(keyed, start=1)]


def compute_key_figures(rows):
    """Recompute every key figure the Model sheet holds, in plain Python.

    Returns an ordered list of (figure, value_string) pairs, in the same
    order the Model sheet lays them out.
    """
    regions, table = pivot(rows)
    figures = []

    def fmt(value):
        return "n/a" if value is None else f"{value}"

    for region in regions:
        cells = table[region]
        s = slug(region)
        for scenario in SCENARIOS:
            figures.append((f"baseline_{short(scenario)}_{s}",
                            fmt(cells[f"{scenario}_{BASELINE}"])))
        for horizon in HORIZONS:
            for scenario in SCENARIOS:
                figures.append((f"{short(scenario)}_{horizon}_{s}",
                                fmt(cells[f"{scenario}_{horizon}"])))
        for horizon in HORIZONS:
            for scenario in SCENARIOS:
                figures.append((f"delta_{short(scenario)}_{horizon}_{s}",
                                fmt(cells[f"delta_{scenario}_{horizon}"])))

    ranking = ranked_regions(regions, table)
    for rank, region, _value in ranking:
        figures.append((f"rank_2095_rcp85_{slug(region)}", str(rank)))

    _top_rank, top_region, top_delta = ranking[0]
    d45 = [table[r]["delta_RCP4.5_2095"] for r in regions
           if table[r]["delta_RCP4.5_2095"] is not None]
    d85 = [table[r]["delta_RCP8.5_2095"] for r in regions
           if table[r]["delta_RCP8.5_2095"] is not None]
    avg45 = two_dp(sum(d45) / len(d45))
    avg85 = two_dp(sum(d85) / len(d85))
    figures.append(("top_region_2095_rcp85", top_region))
    figures.append(("top_delta_2095_rcp85", f"{top_delta}"))
    figures.append(("avg_delta_rcp45_2095", f"{avg45}"))
    figures.append(("avg_delta_rcp85_2095", f"{avg85}"))
    figures.append(("scenario_gap_2095", f"{two_dp(avg85 - avg45)}"))
    return figures


def build_workbook(rows):
    """Write the .xlsx: a Data sheet of canonical rows and a Model sheet
    where every key figure is a live formula over the Data sheet."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    bold = Font(bold=True)
    wb = Workbook()

    data = wb.active
    data.title = "Data"
    for col, name in enumerate(
            ["region", "scenario", "period", "value_degc"], start=1):
        data.cell(row=1, column=col, value=name).font = bold
    for i, (region, scenario, period, value) in enumerate(rows, start=2):
        data.cell(row=i, column=1, value=region)
        data.cell(row=i, column=2, value=scenario)
        data.cell(row=i, column=3, value=period)
        data.cell(row=i, column=4, value=value).number_format = "0.00"
    for col, width in zip(range(1, 5), (16, 10, 10, 12)):
        data.column_dimensions[get_column_letter(col)].width = width

    regions, table = pivot(rows)
    n = len(regions)
    start = 4                  # first pivot row on the Model sheet
    end = start + n - 1        # last pivot row

    m = wb.create_sheet("Model")
    m["A1"] = ("Climate-projection summary: annual mean temperature deltas "
               "by county (degC, CMIP5 p50)")
    m["A1"].font = Font(bold=True, size=13)

    headers = ["Region",
               "Baseline RCP4.5 (1981-2010)", "Baseline RCP8.5 (1981-2010)",
               "RCP4.5 2045", "RCP8.5 2045", "RCP4.5 2095", "RCP8.5 2095",
               "Delta RCP4.5 2045", "Delta RCP8.5 2045",
               "Delta RCP4.5 2095", "Delta RCP8.5 2095", "Sort key"]
    for col, name in enumerate(headers, start=1):
        m.cell(row=3, column=col, value=name).font = bold

    def slot_formula(row, scenario, period):
        crit = (f'Data!$A:$A,$A{row},Data!$B:$B,"{scenario}",'
                f'Data!$C:$C,"{period}"')
        return (f'=IF(COUNTIFS({crit})=0,"n/a",'
                f'ROUND(AVERAGEIFS(Data!$D:$D,{crit}),2))')

    # Columns B..G hold the six slot means; H..K subtract each scenario's
    # own baseline: H=D-B, I=E-C, J=F-B, K=G-C.
    slots = [(s, BASELINE) for s in SCENARIOS] + \
            [(s, h) for h in HORIZONS for s in SCENARIOS]
    delta_pairs = [("H", "D", "B"), ("I", "E", "C"),
                   ("J", "F", "B"), ("K", "G", "C")]
    for i, region in enumerate(regions):
        r = start + i
        m.cell(row=r, column=1, value=region)
        for j, (scenario, period) in enumerate(slots):
            m.cell(row=r, column=2 + j,
                   value=slot_formula(r, scenario, period)
                   ).number_format = "0.00"
        for delta_col, proj_col, base_col in delta_pairs:
            m.cell(
                row=r, column=ord(delta_col) - ord("A") + 1,
                value=(f'=IF(OR(ISTEXT({base_col}{r}),ISTEXT({proj_col}{r})),'
                       f'"n/a",ROUND({proj_col}{r}-{base_col}{r},2))')
            ).number_format = "0.00"
        # Tie-safe rank key: the 2095 RCP8.5 delta in hundredths, shifted
        # three digits, plus the reverse row position (see sort_key()).
        m.cell(row=r, column=12,
               value=(f'=IF(ISTEXT($K{r}),"n/a",'
                      f'ROUND($K{r}*100,0)*1000+({end}-ROW()))')
               ).number_format = "0"

    delta85 = f"$K${start}:$K${end}"
    keys = f"$L${start}:$L${end}"
    region_col = f"$A${start}:$A${end}"

    t = end + 2
    m.cell(row=t, column=1,
           value="Counties ranked by 2095 RCP8.5 delta").font = bold
    for col, name in enumerate(["Rank", "Region", "Delta (degC)"], start=1):
        m.cell(row=t + 1, column=col, value=name).font = bold
    ranking = ranked_regions(regions, table)
    for k in range(1, len(ranking) + 1):
        r = t + 1 + k
        m.cell(row=r, column=1, value=k)
        m.cell(row=r, column=2,
               value=(f"=INDEX({region_col},MATCH(LARGE({keys},A{r}),"
                      f"{keys},0))"))
        m.cell(row=r, column=3,
               value=(f"=INDEX({delta85},MATCH(LARGE({keys},A{r}),"
                      f"{keys},0))")).number_format = "0.00"

    h = t + len(ranking) + 3
    m.cell(row=h, column=1, value="Headline").font = bold
    m.cell(row=h + 1, column=1, value="Largest 2095 RCP8.5 delta, region")
    m.cell(row=h + 1, column=2, value=f"=B{t + 2}")
    m.cell(row=h + 2, column=1, value="Largest 2095 RCP8.5 delta (degC)")
    m.cell(row=h + 2, column=2, value=f"=C{t + 2}").number_format = "0.00"
    m.cell(row=h + 3, column=1, value="Average 2095 delta, RCP4.5 (degC)")
    m.cell(row=h + 3, column=2,
           value=f"=ROUND(AVERAGE($J${start}:$J${end}),2)"
           ).number_format = "0.00"
    m.cell(row=h + 4, column=1, value="Average 2095 delta, RCP8.5 (degC)")
    m.cell(row=h + 4, column=2,
           value=f"=ROUND(AVERAGE({delta85}),2)").number_format = "0.00"
    m.cell(row=h + 5, column=1, value="Scenario gap at 2095 (degC)")
    m.cell(row=h + 5, column=2,
           value=f"=ROUND(B{h + 4}-B{h + 3},2)").number_format = "0.00"

    m.column_dimensions["A"].width = 34
    for col in range(2, 13):
        m.column_dimensions[get_column_letter(col)].width = 17

    wb.save(WORKBOOK)


def verify():
    """Recompute every key figure and diff against expected/key_figures.csv."""
    computed = compute_key_figures(prepare_rows(load_raw()))
    if not os.path.exists(EXPECTED):
        sys.exit("expected/key_figures.csv is missing")
    expected = []
    with open(EXPECTED, newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        header = next(reader, None)
        if header != ["figure", "value"]:
            sys.exit("expected/key_figures.csv has an unexpected header")
        for rec in reader:
            if rec:
                expected.append((rec[0], rec[1]))
    if len(computed) != len(expected):
        print(f"FAIL: {len(computed)} computed figures, "
              f"{len(expected)} expected")
        sys.exit(1)
    for (cf, cv), (ef, ev) in zip(computed, expected):
        if cf != ef or cv != ev:
            print(f"FAIL: first mismatch at {ef}: expected {ev}, got "
                  f"{cv} (computed name {cf})")
            sys.exit(1)
    print(f"PASS: all {len(computed)} key figures match "
          "expected/key_figures.csv")


def show():
    """Print the key figures as an aligned plain-ASCII table."""
    computed = compute_key_figures(prepare_rows(load_raw()))
    name_width = max(len(name) for name, _ in computed)
    value_width = max(len(value) for _, value in computed)
    print(f"{'figure'.ljust(name_width)}  {'value'.rjust(value_width)}")
    print(f"{'-' * name_width}  {'-' * value_width}")
    for name, value in computed:
        print(f"{name.ljust(name_width)}  {value.rjust(value_width)}")


def main():
    command = sys.argv[1] if len(sys.argv) > 1 else "build"
    if command == "build":
        rows = prepare_rows(load_raw())
        build_workbook(rows)
        print(f"Wrote {os.path.basename(WORKBOOK)} ({len(rows)} data rows)")
        verify()
    elif command == "verify":
        verify()
    elif command == "show":
        show()
    else:
        sys.exit("Usage: python build.py [verify|show]")


if __name__ == "__main__":
    main()
